import pandas as pd
from io import BytesIO
from tempfile import NamedTemporaryFile
from datetime import datetime
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.dimensions import ColumnDimension, SheetDimension, SheetFormatProperties
from openpyxl.worksheet.pagebreak import Break, PageBreak
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import PatternFill, NamedStyle, Border, Side, Alignment, Protection, Font, Color, colors, numbers, DEFAULT_FONT
from openpyxl import drawing

def comparison_inputs_to_excel(
        agent,
        date,
        offer_qty,
        seller_name,
        seller_address,
        list_price,
        first_trust,
        second_trust,
        annual_taxes,
        prorated_taxes,
        annual_hoa_condo_fees,
        prorated_hoa_condo_fees,
        listing_company_pct,
        selling_company_pct,
        processing_fee,
        settlement_fee,
        deed_preparation_fee,
        lien_trust_release_fee,
        lien_trust_release_qty,
        recording_fee,
        recording_trusts_liens_qty,
        grantors_tax_pct,
        congestion_tax_pct,
        pest_inspection_fee,
        poa_condo_disclosure_fee,
        offer_1_name,
        offer_1_amt,
        offer_1_down_pmt_pct,
        offer_1_settlement_date,
        offer_1_settlement_company,
        offer_1_emd_amt,
        offer_1_financing_type,
        offer_1_home_inspection_check,
        offer_1_home_inspection_days,
        offer_1_radon_inspection_check,
        offer_1_radon_inspection_days,
        offer_1_septic_inspection_check,
        offer_1_septic_inspection_days,
        offer_1_well_inspection_check,
        offer_1_well_inspection_days,
        offer_1_finance_contingency_check,
        offer_1_finance_contingency_days,
        offer_1_appraisal_contingency_check,
        offer_1_appraisal_contingency_days,
        offer_1_home_sale_contingency_check,
        offer_1_home_sale_contingency_days,
        offer_1_pre_occupancy_date,
        offer_1_post_occupancy_date,
        offer_1_closing_cost_subsidy_amt,
        offer_1_pre_occupancy_credit_amt,
        offer_1_post_occupancy_cost_amt,
        offer_2_name,
        offer_2_amt,
        offer_2_down_pmt_pct,
        offer_2_settlement_date,
        offer_2_settlement_company,
        offer_2_emd_amt,
        offer_2_financing_type,
        offer_2_home_inspection_check,
        offer_2_home_inspection_days,
        offer_2_radon_inspection_check,
        offer_2_radon_inspection_days,
        offer_2_septic_inspection_check,
        offer_2_septic_inspection_days,
        offer_2_well_inspection_check,
        offer_2_well_inspection_days,
        offer_2_finance_contingency_check,
        offer_2_finance_contingency_days,
        offer_2_appraisal_contingency_check,
        offer_2_appraisal_contingency_days,
        offer_2_home_sale_contingency_check,
        offer_2_home_sale_contingency_days,
        offer_2_pre_occupancy_date,
        offer_2_post_occupancy_date,
        offer_2_closing_cost_subsidy_amt,
        offer_2_pre_occupancy_credit_amt,
        offer_2_post_occupancy_cost_amt,
        offer_3_name,
        offer_3_amt,
        offer_3_down_pmt_pct,
        offer_3_settlement_date,
        offer_3_settlement_company,
        offer_3_emd_amt,
        offer_3_financing_type,
        offer_3_home_inspection_check,
        offer_3_home_inspection_days,
        offer_3_radon_inspection_check,
        offer_3_radon_inspection_days,
        offer_3_septic_inspection_check,
        offer_3_septic_inspection_days,
        offer_3_well_inspection_check,
        offer_3_well_inspection_days,
        offer_3_finance_contingency_check,
        offer_3_finance_contingency_days,
        offer_3_appraisal_contingency_check,
        offer_3_appraisal_contingency_days,
        offer_3_home_sale_contingency_check,
        offer_3_home_sale_contingency_days,
        offer_3_pre_occupancy_date,
        offer_3_post_occupancy_date,
        offer_3_closing_cost_subsidy_amt,
        offer_3_pre_occupancy_credit_amt,
        offer_3_post_occupancy_cost_amt,
        offer_4_name,
        offer_4_amt,
        offer_4_down_pmt_pct,
        offer_4_settlement_date,
        offer_4_settlement_company,
        offer_4_emd_amt,
        offer_4_financing_type,
        offer_4_home_inspection_check,
        offer_4_home_inspection_days,
        offer_4_radon_inspection_check,
        offer_4_radon_inspection_days,
        offer_4_septic_inspection_check,
        offer_4_septic_inspection_days,
        offer_4_well_inspection_check,
        offer_4_well_inspection_days,
        offer_4_finance_contingency_check,
        offer_4_finance_contingency_days,
        offer_4_appraisal_contingency_check,
        offer_4_appraisal_contingency_days,
        offer_4_home_sale_contingency_check,
        offer_4_home_sale_contingency_days,
        offer_4_pre_occupancy_date,
        offer_4_post_occupancy_date,
        offer_4_closing_cost_subsidy_amt,
        offer_4_pre_occupancy_credit_amt,
        offer_4_post_occupancy_cost_amt,
        offer_5_name,
        offer_5_amt,
        offer_5_down_pmt_pct,
        offer_5_settlement_date,
        offer_5_settlement_company,
        offer_5_emd_amt,
        offer_5_financing_type,
        offer_5_home_inspection_check,
        offer_5_home_inspection_days,
        offer_5_radon_inspection_check,
        offer_5_radon_inspection_days,
        offer_5_septic_inspection_check,
        offer_5_septic_inspection_days,
        offer_5_well_inspection_check,
        offer_5_well_inspection_days,
        offer_5_finance_contingency_check,
        offer_5_finance_contingency_days,
        offer_5_appraisal_contingency_check,
        offer_5_appraisal_contingency_days,
        offer_5_home_sale_contingency_check,
        offer_5_home_sale_contingency_days,
        offer_5_pre_occupancy_date,
        offer_5_post_occupancy_date,
        offer_5_closing_cost_subsidy_amt,
        offer_5_pre_occupancy_credit_amt,
        offer_5_post_occupancy_cost_amt,
        offer_6_name,
        offer_6_amt,
        offer_6_down_pmt_pct,
        offer_6_settlement_date,
        offer_6_settlement_company,
        offer_6_emd_amt,
        offer_6_financing_type,
        offer_6_home_inspection_check,
        offer_6_home_inspection_days,
        offer_6_radon_inspection_check,
        offer_6_radon_inspection_days,
        offer_6_septic_inspection_check,
        offer_6_septic_inspection_days,
        offer_6_well_inspection_check,
        offer_6_well_inspection_days,
        offer_6_finance_contingency_check,
        offer_6_finance_contingency_days,
        offer_6_appraisal_contingency_check,
        offer_6_appraisal_contingency_days,
        offer_6_home_sale_contingency_check,
        offer_6_home_sale_contingency_days,
        offer_6_pre_occupancy_date,
        offer_6_post_occupancy_date,
        offer_6_closing_cost_subsidy_amt,
        offer_6_pre_occupancy_credit_amt,
        offer_6_post_occupancy_cost_amt,
        offer_7_name,
        offer_7_amt,
        offer_7_down_pmt_pct,
        offer_7_settlement_date,
        offer_7_settlement_company,
        offer_7_emd_amt,
        offer_7_financing_type,
        offer_7_home_inspection_check,
        offer_7_home_inspection_days,
        offer_7_radon_inspection_check,
        offer_7_radon_inspection_days,
        offer_7_septic_inspection_check,
        offer_7_septic_inspection_days,
        offer_7_well_inspection_check,
        offer_7_well_inspection_days,
        offer_7_finance_contingency_check,
        offer_7_finance_contingency_days,
        offer_7_appraisal_contingency_check,
        offer_7_appraisal_contingency_days,
        offer_7_home_sale_contingency_check,
        offer_7_home_sale_contingency_days,
        offer_7_pre_occupancy_date,
        offer_7_post_occupancy_date,
        offer_7_closing_cost_subsidy_amt,
        offer_7_pre_occupancy_credit_amt,
        offer_7_post_occupancy_cost_amt,
        offer_8_name,
        offer_8_amt,
        offer_8_down_pmt_pct,
        offer_8_settlement_date,
        offer_8_settlement_company,
        offer_8_emd_amt,
        offer_8_financing_type,
        offer_8_home_inspection_check,
        offer_8_home_inspection_days,
        offer_8_radon_inspection_check,
        offer_8_radon_inspection_days,
        offer_8_septic_inspection_check,
        offer_8_septic_inspection_days,
        offer_8_well_inspection_check,
        offer_8_well_inspection_days,
        offer_8_finance_contingency_check,
        offer_8_finance_contingency_days,
        offer_8_appraisal_contingency_check,
        offer_8_appraisal_contingency_days,
        offer_8_home_sale_contingency_check,
        offer_8_home_sale_contingency_days,
        offer_8_pre_occupancy_date,
        offer_8_post_occupancy_date,
        offer_8_closing_cost_subsidy_amt,
        offer_8_pre_occupancy_credit_amt,
        offer_8_post_occupancy_cost_amt,
        offer_9_name,
        offer_9_amt,
        offer_9_down_pmt_pct,
        offer_9_settlement_date,
        offer_9_settlement_company,
        offer_9_emd_amt,
        offer_9_financing_type,
        offer_9_home_inspection_check,
        offer_9_home_inspection_days,
        offer_9_radon_inspection_check,
        offer_9_radon_inspection_days,
        offer_9_septic_inspection_check,
        offer_9_septic_inspection_days,
        offer_9_well_inspection_check,
        offer_9_well_inspection_days,
        offer_9_finance_contingency_check,
        offer_9_finance_contingency_days,
        offer_9_appraisal_contingency_check,
        offer_9_appraisal_contingency_days,
        offer_9_home_sale_contingency_check,
        offer_9_home_sale_contingency_days,
        offer_9_pre_occupancy_date,
        offer_9_post_occupancy_date,
        offer_9_closing_cost_subsidy_amt,
        offer_9_pre_occupancy_credit_amt,
        offer_9_post_occupancy_cost_amt,
        offer_10_name,
        offer_10_amt,
        offer_10_down_pmt_pct,
        offer_10_settlement_date,
        offer_10_settlement_company,
        offer_10_emd_amt,
        offer_10_financing_type,
        offer_10_home_inspection_check,
        offer_10_home_inspection_days,
        offer_10_radon_inspection_check,
        offer_10_radon_inspection_days,
        offer_10_septic_inspection_check,
        offer_10_septic_inspection_days,
        offer_10_well_inspection_check,
        offer_10_well_inspection_days,
        offer_10_finance_contingency_check,
        offer_10_finance_contingency_days,
        offer_10_appraisal_contingency_check,
        offer_10_appraisal_contingency_days,
        offer_10_home_sale_contingency_check,
        offer_10_home_sale_contingency_days,
        offer_10_pre_occupancy_date,
        offer_10_post_occupancy_date,
        offer_10_closing_cost_subsidy_amt,
        offer_10_pre_occupancy_credit_amt,
        offer_10_post_occupancy_cost_amt,
        offer_11_name,
        offer_11_amt,
        offer_11_down_pmt_pct,
        offer_11_settlement_date,
        offer_11_settlement_company,
        offer_11_emd_amt,
        offer_11_financing_type,
        offer_11_home_inspection_check,
        offer_11_home_inspection_days,
        offer_11_radon_inspection_check,
        offer_11_radon_inspection_days,
        offer_11_septic_inspection_check,
        offer_11_septic_inspection_days,
        offer_11_well_inspection_check,
        offer_11_well_inspection_days,
        offer_11_finance_contingency_check,
        offer_11_finance_contingency_days,
        offer_11_appraisal_contingency_check,
        offer_11_appraisal_contingency_days,
        offer_11_home_sale_contingency_check,
        offer_11_home_sale_contingency_days,
        offer_11_pre_occupancy_date,
        offer_11_post_occupancy_date,
        offer_11_closing_cost_subsidy_amt,
        offer_11_pre_occupancy_credit_amt,
        offer_11_post_occupancy_cost_amt,
        offer_12_name,
        offer_12_amt,
        offer_12_down_pmt_pct,
        offer_12_settlement_date,
        offer_12_settlement_company,
        offer_12_emd_amt,
        offer_12_financing_type,
        offer_12_home_inspection_check,
        offer_12_home_inspection_days,
        offer_12_radon_inspection_check,
        offer_12_radon_inspection_days,
        offer_12_septic_inspection_check,
        offer_12_septic_inspection_days,
        offer_12_well_inspection_check,
        offer_12_well_inspection_days,
        offer_12_finance_contingency_check,
        offer_12_finance_contingency_days,
        offer_12_appraisal_contingency_check,
        offer_12_appraisal_contingency_days,
        offer_12_home_sale_contingency_check,
        offer_12_home_sale_contingency_days,
        offer_12_pre_occupancy_date,
        offer_12_post_occupancy_date,
        offer_12_closing_cost_subsidy_amt,
        offer_12_pre_occupancy_credit_amt,
        offer_12_post_occupancy_cost_amt,
):

    wb = Workbook()
    dest_filename = f"offer_comparison_form_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    ws1 = wb.active
    ws1.title = 'offer_comparison_pg_1'
    ws1.print_area = 'B2:I75'
    ws1.set_printer_settings(paper_size=1, orientation='portrait')

    ws1.page_margins.top = 0.5
    ws1.page_margins.bottom = 0.5
    ws1.page_margins.left = 0.5
    ws1.page_margins.right = 0.5
    ws1.sheet_properties.pageSetUpPr.fitToPage = True
    ws1.print_options.horizontalCentered = True
    ws1.print_options.verticalCentered = True

    ws2 = wb.create_sheet(title='offer_comparison_pg_2')
    ws2.print_area = 'B2:I75'
    ws2.set_printer_settings(paper_size=1, orientation='portrait')

    ws2.page_margins.top = 0.5
    ws2.page_margins.bottom = 0.5
    ws2.page_margins.left = 0.5
    ws2.page_margins.right = 0.5
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.print_options.horizontalCentered = True
    ws2.print_options.verticalCentered = True

    ws3 = wb.create_sheet(title='offer_comparison_pg_3')
    ws3.print_area = 'B2:I75'
    ws3.set_printer_settings(paper_size=1, orientation='portrait')

    ws3.page_margins.top = 0.5
    ws3.page_margins.bottom = 0.5
    ws3.page_margins.left = 0.5
    ws3.page_margins.right = 0.5
    ws3.sheet_properties.pageSetUpPr.fitToPage = True
    ws3.print_options.horizontalCentered = True
    ws3.print_options.verticalCentered = True


    white_fill = '00FFFFFF'
    yellow_fill = '00FFFF00'
    black_fill = '00000000'
    font_size = 12
    thick = Side(border_style='thick')
    thin = Side(border_style='thin')
    hair = Side(border_style='hair')
    DEFAULT_FONT.size = font_size
    ws1.column_dimensions['B'].width = 1.5
    ws1.column_dimensions['C'].width = 40.83
    ws1.column_dimensions['D'].width = 40.83
    ws1.column_dimensions['E'].width = 20.83
    ws1.column_dimensions['F'].width = 20.83
    ws1.column_dimensions['G'].width = 20.83
    ws1.column_dimensions['H'].width = 20.83
    ws1.column_dimensions['I'].width = 1.5

    ws2.column_dimensions['B'].width = 1.5
    ws2.column_dimensions['C'].width = 40.83
    ws2.column_dimensions['D'].width = 40.83
    ws2.column_dimensions['E'].width = 20.83
    ws2.column_dimensions['F'].width = 20.83
    ws2.column_dimensions['G'].width = 20.83
    ws2.column_dimensions['H'].width = 20.83
    ws2.column_dimensions['I'].width = 1.5

    ws3.column_dimensions['B'].width = 1.5
    ws3.column_dimensions['C'].width = 40.83
    ws3.column_dimensions['D'].width = 40.83
    ws3.column_dimensions['E'].width = 20.83
    ws3.column_dimensions['F'].width = 20.83
    ws3.column_dimensions['G'].width = 20.83
    ws3.column_dimensions['H'].width = 20.83
    ws3.column_dimensions['I'].width = 1.5

    acct_fmt = '_($* #,##0_);[Red]_($* (#,##0);_($* "-"??_);_(@_)'
    pct_fmt = '0.00%'
    # date_fmt = NamedStyle(name='date', number_format='DD/MM/YYYY')

    for row in ws1.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
        for cell in row:
            cell.fill = PatternFill(start_color=white_fill, end_color=white_fill, fill_type='solid')

    for row in ws2.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
        for cell in row:
            cell.fill = PatternFill(start_color=white_fill, end_color=white_fill, fill_type='solid')

    for row in ws3.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
        for cell in row:
            cell.fill = PatternFill(start_color=white_fill, end_color=white_fill, fill_type='solid')

    # Build Black Border for pages 1, 2, and 3
    ws1.merge_cells('A1:J1')
    top_left_border_one_1 = ws1['A1']
    top_left_border_one_1.fill = PatternFill('solid', fgColor=black_fill)

    ws1.merge_cells('A2:A76')
    top_left_border_two_1 = ws1['A2']
    top_left_border_two_1.fill = PatternFill('solid', fgColor=black_fill)

    ws1.merge_cells('B76:J76')
    top_left_border_three_1 = ws1['B76']
    top_left_border_three_1.fill = PatternFill('solid', fgColor=black_fill)

    ws1.merge_cells('J2:J75')
    top_left_border_four_1 = ws1['J2']
    top_left_border_four_1.fill = PatternFill('solid', fgColor=black_fill)

    ws2.merge_cells('A1:J1')
    top_left_border_one_2 = ws2['A1']
    top_left_border_one_2.fill = PatternFill('solid', fgColor=black_fill)

    ws2.merge_cells('A2:A76')
    top_left_border_two_2 = ws2['A2']
    top_left_border_two_2.fill = PatternFill('solid', fgColor=black_fill)

    ws2.merge_cells('B76:J76')
    top_left_border_three_2 = ws2['B76']
    top_left_border_three_2.fill = PatternFill('solid', fgColor=black_fill)

    ws2.merge_cells('J2:J75')
    top_left_border_four_2 = ws2['J2']
    top_left_border_four_2.fill = PatternFill('solid', fgColor=black_fill)

    ws3.merge_cells('A1:J1')
    top_left_border_one_3 = ws3['A1']
    top_left_border_one_3.fill = PatternFill('solid', fgColor=black_fill)

    ws3.merge_cells('A2:A76')
    top_left_border_two_3 = ws3['A2']
    top_left_border_two_3.fill = PatternFill('solid', fgColor=black_fill)

    ws3.merge_cells('B76:J76')
    top_left_border_three_3 = ws3['B76']
    top_left_border_three_3.fill = PatternFill('solid', fgColor=black_fill)

    ws3.merge_cells('J2:J75')
    top_left_border_four_3 = ws3['J2']
    top_left_border_four_3.fill = PatternFill('solid', fgColor=black_fill)

    # page 1
    # Build Header
    ws1.merge_cells('C2:H2')
    top_left_cell_one_1 = ws1['C2']
    top_left_cell_one_1.value = 'Seller\'s Total Net Proceeds For Different Offers'
    top_left_cell_one_1.font = Font(bold=True)
    top_left_cell_one_1.alignment = Alignment(horizontal='center')

    ws1.merge_cells('C3:H3')
    top_left_cell_two_1 = ws1['C3']
    top_left_cell_two_1.value = f'{seller_name} - {seller_address}'
    top_left_cell_two_1.font = Font(bold=True)
    top_left_cell_two_1.alignment = Alignment(horizontal='center')

    ws1.merge_cells('C4:H4')
    top_left_cell_three_1 = ws1['C4']
    top_left_cell_three_1.value = f'Date Prepared: {date}'
    top_left_cell_three_1.font = Font(bold=True)
    top_left_cell_three_1.alignment = Alignment(horizontal='center')

    ws1.merge_cells('C5:H5')
    top_left_cell_four_1 = ws1['C5']
    top_left_cell_four_1.value = f'List Price: ${list_price:,.2f}'
    top_left_cell_four_1.font = Font(bold=True)
    top_left_cell_four_1.alignment = Alignment(horizontal='center')

    c7_1 = ws1['C7']
    c7_1.value = 'OFFER SUMMARY FEATURES'
    c7_1.font = Font(bold=True)
    c7_1.border = Border(bottom=thin)

    e7_1 = ws1['E7']
    e7_1.value = offer_1_name
    e7_1.font = Font(bold=True)
    e7_1.border = Border(bottom=thin)
    e7_1.alignment = Alignment(horizontal='center', wrap_text=True)

    f7_1 = ws1['F7']
    f7_1.value = offer_2_name
    f7_1.font = Font(bold=True)
    f7_1.border = Border(bottom=thin)
    f7_1.alignment = Alignment(horizontal='center', wrap_text=True)

    g7_1 = ws1['G7']
    g7_1.value = offer_3_name
    g7_1.font = Font(bold=True)
    g7_1.border = Border(bottom=thin)
    g7_1.alignment = Alignment(horizontal='center', wrap_text=True)

    h7_1 = ws1['H7']
    h7_1.value = offer_4_name
    h7_1.font = Font(bold=True)
    h7_1.border = Border(bottom=thin)
    h7_1.alignment = Alignment(horizontal='center', wrap_text=True)

    c8_1 = ws1['C8']
    c8_1.value = 'Offer Amt. ($)'
    # c8_1.font = Font(bold=True)
    c8_1.alignment = Alignment(horizontal='left', vertical='center')
    c8_1.border = Border(top=thin, bottom=hair, left=thin)

    d8_1 = ws1['D8']
    d8_1.border = Border(top=thin, bottom=hair)

    e8_1 = ws1['E8']
    e8_1.value = offer_1_amt
    # e8_1.font = Font(bold=True)
    e8_1.number_format = acct_fmt
    e8_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f8_1 = ws1['F8']
    f8_1.value = offer_2_amt
    # f8_1.font = Font(bold=True)
    f8_1.number_format = acct_fmt
    f8_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g8_1 = ws1['G8']
    g8_1.value = offer_3_amt
    # g8_1.font = Font(bold=True)
    g8_1.number_format = acct_fmt
    g8_1.border = Border(top=thin, bottom=hair, right=thin, left=thin)

    h8_1 = ws1['H8']
    h8_1.value = offer_4_amt
    # h8_1.font = Font(bold=True)
    h8_1.number_format = acct_fmt
    h8_1.border = Border(top=thin, bottom=hair, right=thin, left=thin)

    c9_1 = ws1['C9']
    c9_1.value = 'Down Pmt (%)'
    # c9_1.font = Font(bold=True)
    c9_1.alignment = Alignment(horizontal='left', vertical='center')
    c9_1.border = Border(top=hair, bottom=hair, left=thin)

    d9_1 = ws1['D9']
    d9_1.border = Border(top=hair, bottom=hair)

    e9_1 = ws1['E9']
    e9_1.value = offer_1_down_pmt_pct
    # e9_1.font = Font(bold=True)
    e9_1.number_format = pct_fmt
    e9_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f9_1 = ws1['F9']
    f9_1.value = offer_2_down_pmt_pct
    # f9_1.font = Font(bold=True)
    f9_1.number_format = pct_fmt
    f9_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g9_1 = ws1['G9']
    g9_1.value = offer_3_down_pmt_pct
    # g9_1.font = Font(bold=True)
    g9_1.number_format = pct_fmt
    g9_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h9_1 = ws1['H9']
    h9_1.value = offer_4_down_pmt_pct
    # h9_1.font = Font(bold=True)
    h9_1.number_format = pct_fmt
    h9_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c10_1 = ws1['C10']
    c10_1.value = 'Settlement Date'
    # c10_1.font = Font(bold=True)
    c10_1.alignment = Alignment(horizontal='left', vertical='center')
    c10_1.border = Border(top=hair, bottom=hair, left=thin)

    d10_1 = ws1['D10']
    d10_1.border = Border(top=hair, bottom=hair)

    e10_1 = ws1['E10']
    e10_1.value = offer_1_settlement_date
    # e10_1.font = Font(bold=True)
    e10_1.alignment = Alignment(horizontal='right')
    e10_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f10_1 = ws1['F10']
    f10_1.value = offer_2_settlement_date
    # f10_1.font = Font(bold=True)
    f10_1.alignment = Alignment(horizontal='right')
    f10_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g10_1 = ws1['G10']
    g10_1.value = offer_3_settlement_date
    # g10_1.font = Font(bold=True)
    g10_1.alignment = Alignment(horizontal='right')
    g10_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h10_1 = ws1['H10']
    h10_1.value = offer_4_settlement_date
    # h10_1.font = Font(bold=True)
    h10_1.alignment = Alignment(horizontal='right')
    h10_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c11_1 = ws1['C11']
    c11_1.value = 'Settlement Company'
    # c11_1.font = Font(bold=True)
    c11_1.alignment = Alignment(horizontal='left', vertical='center')
    c11_1.border = Border(top=hair, bottom=hair, left=thin)

    d11_1 = ws1['D11']
    d11_1.border = Border(top=hair, bottom=hair)

    e11_1 = ws1['E11']
    e11_1.value = offer_1_settlement_company
    # e11_1.font = Font(bold=True)
    e11_1.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    e11_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f11_1 = ws1['F11']
    f11_1.value = offer_2_settlement_company
    # f11_1.font = Font(bold=True)
    f11_1.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    f11_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g11_1 = ws1['G11']
    g11_1.value = offer_3_settlement_company
    # g11_1.font = Font(bold=True)
    g11_1.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    g11_1.border = Border(top=hair, bottom=hair, right=thin, left=thin)

    h11_1 = ws1['H11']
    h11_1.value = offer_4_settlement_company
    # h11_1.font = Font(bold=True)
    h11_1.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    h11_1.border = Border(top=hair, bottom=hair, right=thin, left=thin)

    c12_1 = ws1['C12']
    c12_1.value = 'EMD Amt. ($)'
    # c12_1.font = Font(bold=True)
    c12_1.alignment = Alignment(horizontal='left', vertical='center')
    c12_1.border = Border(top=hair, bottom=hair, left=thin)

    d12_1 = ws1['D12']
    d12_1.border = Border(top=hair, bottom=hair)

    e12_1 = ws1['E12']
    e12_1.value = offer_1_emd_amt
    # e12_1.font = Font(bold=True)
    e12_1.number_format = acct_fmt
    e12_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f12_1 = ws1['F12']
    f12_1.value = offer_2_emd_amt
    # f12_1.font = Font(bold=True)
    f12_1.number_format = acct_fmt
    f12_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g12_1 = ws1['G12']
    g12_1.value = offer_3_emd_amt
    # g12_1.font = Font(bold=True)
    g12_1.number_format = acct_fmt
    g12_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h12_1 = ws1['H12']
    h12_1.value = offer_4_emd_amt
    # h12_1.font = Font(bold=True)
    h12_1.number_format = acct_fmt
    h12_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c13_1 = ws1['C13']
    c13_1.value = 'Financing Type'
    # c13_1.font = Font(bold=True)
    c13_1.alignment = Alignment(horizontal='left', vertical='center')
    c13_1.border = Border(top=hair, bottom=hair, left=thin)

    d13_1 = ws1['D13']
    d13_1.border = Border(top=hair, bottom=hair)

    e13_1 = ws1['E13']
    e13_1.value = offer_1_financing_type
    # e13_1.font = Font(bold=True)
    e13_1.alignment = Alignment(horizontal='center')
    e13_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f13_1 = ws1['F13']
    f13_1.value = offer_2_financing_type
    # f13_1.font = Font(bold=True)
    f13_1.alignment = Alignment(horizontal='center')
    f13_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g13_1 = ws1['G13']
    g13_1.value = offer_3_financing_type
    # g13_1.font = Font(bold=True)
    g13_1.alignment = Alignment(horizontal='center')
    g13_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h13_1 = ws1['H13']
    h13_1.value = offer_4_financing_type
    # h13_1.font = Font(bold=True)
    h13_1.alignment = Alignment(horizontal='center')
    h13_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    ws1.merge_cells('C14:C15')
    top_left_home_insp_cont_1 = ws1['C14']
    top_left_home_insp_cont_1.value = 'Home Inspection Contingency'
    # top_left_home_insp_cont_1.font = Font(bold=True)
    top_left_home_insp_cont_1.alignment = Alignment(horizontal='left', vertical='center')
    top_left_home_insp_cont_1.border = Border(top=hair, bottom=hair, left=thin)
    c14_1 = ws1['C14']
    c14_1.border = Border(top=hair, left=thin)
    d14_1 = ws1['D14']
    d14_1.border = Border(top=hair)
    c15_1 = ws1['C15']
    c15_1.border = Border(bottom=hair, left=thin)
    d15_1 = ws1['D15']
    d15_1.border = Border(bottom=hair)

    e14_1 = ws1['E14']
    e14_1.value = offer_1_home_inspection_check
    # e14_1.font = Font(bold=True)
    e14_1.alignment = Alignment(horizontal='center')
    e14_1.border = Border(top=hair, left=thin, right=thin)

    f14_1 = ws1['F14']
    f14_1.value = offer_2_home_inspection_check
    # f14_1.font = Font(bold=True)
    f14_1.alignment = Alignment(horizontal='center')
    f14_1.border = Border(top=hair, left=thin, right=thin)

    g14_1 = ws1['G14']
    g14_1.value = offer_3_home_inspection_check
    # g14_1.font = Font(bold=True)
    g14_1.alignment = Alignment(horizontal='center')
    g14_1.border = Border(top=hair, left=thin, right=thin)

    h14_1 = ws1['H14']
    h14_1.value = offer_4_home_inspection_check
    # h14_1.font = Font(bold=True)
    h14_1.alignment = Alignment(horizontal='center')
    h14_1.border = Border(top=hair, left=thin, right=thin)

    e15_1 = ws1['E15']
    e15_1.value = offer_1_home_inspection_days
    # e15_1.font = Font(bold=True)
    e15_1.alignment = Alignment(horizontal='center')
    e15_1.border = Border(bottom=hair, left=thin, right=thin)

    f15_1 = ws1['F15']
    f15_1.value = offer_2_home_inspection_days
    # f15_1.font = Font(bold=True)
    f15_1.alignment = Alignment(horizontal='center')
    f15_1.border = Border(bottom=hair, left=thin, right=thin)

    g15_1 = ws1['G15']
    g15_1.value = offer_3_home_inspection_days
    # g15_1.font = Font(bold=True)
    g15_1.alignment = Alignment(horizontal='center')
    g15_1.border = Border(bottom=hair, left=thin, right=thin)

    h15_1 = ws1['H15']
    h15_1.value = offer_4_home_inspection_days
    # h15_1.font = Font(bold=True)
    h15_1.alignment = Alignment(horizontal='center')
    h15_1.border = Border(bottom=hair, left=thin, right=thin)

    ws1.merge_cells('C16:C17')
    top_left_radon_insp_cont_1 = ws1['C16']
    top_left_radon_insp_cont_1.value = 'Radon Inspection Contingency'
    # top_left_radon_insp_cont_1.font = Font(bold=True)
    top_left_radon_insp_cont_1.alignment = Alignment(horizontal='left', vertical='center')
    top_left_radon_insp_cont_1.border = Border(top=hair, bottom=hair, left=thin)
    c16_1 = ws1['C16']
    c16_1.border = Border(top=hair, left=thin)
    d16_1 = ws1['D16']
    d16_1.border = Border(top=hair)
    c17_1 = ws1['C17']
    c17_1.border = Border(bottom=hair, left=thin)
    d17_1 = ws1['D17']
    d17_1.border = Border(bottom=hair)

    e16_1 = ws1['E16']
    e16_1.value = offer_1_radon_inspection_check
    # e16_1.font = Font(bold=True)
    e16_1.alignment = Alignment(horizontal='center')
    e16_1.border = Border(top=hair, left=thin, right=thin)

    f16_1 = ws1['F16']
    f16_1.value = offer_2_radon_inspection_check
    # f16_1.font = Font(bold=True)
    f16_1.alignment = Alignment(horizontal='center')
    f16_1.border = Border(top=hair, left=thin, right=thin)

    g16_1 = ws1['G16']
    g16_1.value = offer_3_radon_inspection_check
    # g16_1.font = Font(bold=True)
    g16_1.alignment = Alignment(horizontal='center')
    g16_1.border = Border(top=hair, left=thin, right=thin)

    h16_1 = ws1['H16']
    h16_1.value = offer_4_radon_inspection_check
    # h16_1.font = Font(bold=True)
    h16_1.alignment = Alignment(horizontal='center')
    h16_1.border = Border(top=hair, left=thin, right=thin)

    e17_1 = ws1['E17']
    e17_1.value = offer_1_radon_inspection_days
    # e17_1.font = Font(bold=True)
    e17_1.alignment = Alignment(horizontal='center')
    e17_1.border = Border(bottom=hair, left=thin, right=thin)

    f17_1 = ws1['F17']
    f17_1.value = offer_2_radon_inspection_days
    # f17_1.font = Font(bold=True)
    f17_1.alignment = Alignment(horizontal='center')
    f17_1.border = Border(bottom=hair, left=thin, right=thin)

    g17_1 = ws1['G17']
    g17_1.value = offer_3_radon_inspection_days
    # g17_1.font = Font(bold=True)
    g17_1.alignment = Alignment(horizontal='center')
    g17_1.border = Border(bottom=hair, left=thin, right=thin)

    h17_1 = ws1['H17']
    h17_1.value = offer_4_radon_inspection_days
    # h17_1.font = Font(bold=True)
    h17_1.alignment = Alignment(horizontal='center')
    h17_1.border = Border(bottom=hair, left=thin, right=thin)

    ws1.merge_cells('C18:C19')
    top_left_septic_insp_cont_1 = ws1['C18']
    top_left_septic_insp_cont_1.value = 'Septic Inspection Contingency'
    # top_left_septic_insp_cont_1.font = Font(bold=True)
    top_left_septic_insp_cont_1.alignment = Alignment(horizontal='left', vertical='center')
    top_left_septic_insp_cont_1.border = Border(top=hair, bottom=hair, left=thin)
    c18_1 = ws1['C18']
    c18_1.border = Border(top=hair, left=thin)
    d18_1 = ws1['D18']
    d18_1.border = Border(top=hair)
    c19_1 = ws1['C19']
    c19_1.border = Border(bottom=hair, left=thin)
    d19_1 = ws1['D19']
    d19_1.border = Border(bottom=hair)

    e18_1 = ws1['E18']
    e18_1.value = offer_1_septic_inspection_check
    # e18_1.font = Font(bold=True)
    e18_1.alignment = Alignment(horizontal='center')
    e18_1.border = Border(top=hair, left=thin, right=thin)

    f18_1 = ws1['F18']
    f18_1.value = offer_2_septic_inspection_check
    # f18_1.font = Font(bold=True)
    f18_1.alignment = Alignment(horizontal='center')
    f18_1.border = Border(top=hair, left=thin, right=thin)

    g18_1 = ws1['G18']
    g18_1.value = offer_3_septic_inspection_check
    # g18_1.font = Font(bold=True)
    g18_1.alignment = Alignment(horizontal='center')
    g18_1.border = Border(top=hair, left=thin, right=thin)

    h18_1 = ws1['H18']
    h18_1.value = offer_4_septic_inspection_check
    # h18_1.font = Font(bold=True)
    h18_1.alignment = Alignment(horizontal='center')
    h18_1.border = Border(top=hair, left=thin, right=thin)

    e19_1 = ws1['E19']
    e19_1.value = offer_1_septic_inspection_days
    # e19_1.font = Font(bold=True)
    e19_1.alignment = Alignment(horizontal='center')
    e19_1.border = Border(bottom=hair, left=thin, right=thin)

    f19_1 = ws1['F19']
    f19_1.value = offer_2_septic_inspection_days
    # f19_1.font = Font(bold=True)
    f19_1.alignment = Alignment(horizontal='center')
    f19_1.border = Border(bottom=hair, left=thin, right=thin)

    g19_1 = ws1['G19']
    g19_1.value = offer_3_septic_inspection_days
    # g19_1.font = Font(bold=True)
    g19_1.alignment = Alignment(horizontal='center')
    g19_1.border = Border(bottom=hair, left=thin, right=thin)

    h19_1 = ws1['H19']
    h19_1.value = offer_4_septic_inspection_days
    # h19_1.font = Font(bold=True)
    h19_1.alignment = Alignment(horizontal='center')
    h19_1.border = Border(bottom=hair, left=thin, right=thin)

    ws1.merge_cells('C20:C21')
    top_left_well_insp_cont_1 = ws1['C20']
    top_left_well_insp_cont_1.value = 'Well Inspection Contingency'
    # top_left_well_insp_cont_1.font = Font(bold=True)
    top_left_well_insp_cont_1.alignment = Alignment(horizontal='left', vertical='center')
    top_left_well_insp_cont_1.border = Border(top=hair, bottom=hair, left=thin)
    c20_1 = ws1['C20']
    c20_1.border = Border(top=hair, left=thin)
    d20_1 = ws1['D20']
    d20_1.border = Border(top=hair)
    c21_1 = ws1['C21']
    c21_1.border = Border(bottom=hair, left=thin)
    d21_1 = ws1['D21']
    d21_1.border = Border(bottom=hair)

    e20_1 = ws1['E20']
    e20_1.value = offer_1_well_inspection_check
    # e20_1.font = Font(bold=True)
    e20_1.alignment = Alignment(horizontal='center')
    e20_1.border = Border(top=hair, left=thin, right=thin)

    f20_1 = ws1['F20']
    f20_1.value = offer_2_well_inspection_check
    # f20_1.font = Font(bold=True)
    f20_1.alignment = Alignment(horizontal='center')
    f20_1.border = Border(top=hair, left=thin, right=thin)

    g20_1 = ws1['G20']
    g20_1.value = offer_3_well_inspection_check
    # g20_1.font = Font(bold=True)
    g20_1.alignment = Alignment(horizontal='center')
    g20_1.border = Border(top=hair, left=thin, right=thin)

    h20_1 = ws1['H20']
    h20_1.value = offer_4_well_inspection_check
    # h20_1.font = Font(bold=True)
    h20_1.alignment = Alignment(horizontal='center')
    h20_1.border = Border(top=hair, left=thin, right=thin)

    e21_1 = ws1['E21']
    e21_1.value = offer_1_well_inspection_days
    # e21_1.font = Font(bold=True)
    e21_1.alignment = Alignment(horizontal='center')
    e21_1.border = Border(bottom=hair, left=thin, right=thin)

    f21_1 = ws1['F21']
    f21_1.value = offer_2_well_inspection_days
    # f21_1.font = Font(bold=True)
    f21_1.alignment = Alignment(horizontal='center')
    f21_1.border = Border(bottom=hair, left=thin, right=thin)

    g21_1 = ws1['G21']
    g21_1.value = offer_3_well_inspection_days
    # g21_1.font = Font(bold=True)
    g21_1.alignment = Alignment(horizontal='center')
    g21_1.border = Border( bottom=hair, left=thin, right=thin)

    h21_1 = ws1['H21']
    h21_1.value = offer_4_well_inspection_days
    # h21_1.font = Font(bold=True)
    h21_1.alignment = Alignment(horizontal='center')
    h21_1.border = Border(bottom=hair, left=thin, right=thin)

    ws1.merge_cells('C22:C23')
    top_left_finance_cont_1 = ws1['C22']
    top_left_finance_cont_1.value = 'Finance Contingency'
    # top_left_finance_cont_1.font = Font(bold=True)
    top_left_finance_cont_1.alignment = Alignment(horizontal='left', vertical='center')
    top_left_finance_cont_1.border = Border(top=hair, bottom=hair, left=thin)
    c22_1 = ws1['C22']
    c22_1.border = Border(top=hair, left=thin)
    d22_1 = ws1['D22']
    d22_1.border = Border(top=hair)
    c23_1 = ws1['C23']
    c23_1.border = Border(bottom=hair, left=thin)
    d23_1 = ws1['D23']
    d23_1.border = Border(bottom=hair)

    e22_1 = ws1['E22']
    e22_1.value = offer_1_finance_contingency_check
    # e22_1.font = Font(bold=True)
    e22_1.alignment = Alignment(horizontal='center')
    e22_1.border = Border(top=hair, left=thin, right=thin)

    f22_1 = ws1['F22']
    f22_1.value = offer_2_finance_contingency_check
    # f22_1.font = Font(bold=True)
    f22_1.alignment = Alignment(horizontal='center')
    f22_1.border = Border(top=hair, left=thin, right=thin)

    g22_1 = ws1['G22']
    g22_1.value = offer_3_finance_contingency_check
    # g22_1.font = Font(bold=True)
    g22_1.alignment = Alignment(horizontal='center')
    g22_1.border = Border(top=hair, left=thin, right=thin)

    h22_1 = ws1['H22']
    h22_1.value = offer_4_finance_contingency_check
    # h22_1.font = Font(bold=True)
    h22_1.alignment = Alignment(horizontal='center')
    h22_1.border = Border(top=hair, left=thin, right=thin)

    e23_1 = ws1['E23']
    e23_1.value = offer_1_finance_contingency_days
    # e23_1.font = Font(bold=True)
    e23_1.alignment = Alignment(horizontal='center')
    e23_1.border = Border(bottom=hair, left=thin, right=thin)

    f23_1 = ws1['F23']
    f23_1.value = offer_2_finance_contingency_days
    # f23_1.font = Font(bold=True)
    f23_1.alignment = Alignment(horizontal='center')
    f23_1.border = Border(bottom=hair, left=thin, right=thin)

    g23_1 = ws1['G23']
    g23_1.value = offer_3_finance_contingency_days
    # g23_1.font = Font(bold=True)
    g23_1.alignment = Alignment(horizontal='center')
    g23_1.border = Border(bottom=hair, left=thin, right=thin)

    h23_1 = ws1['H23']
    h23_1.value = offer_4_finance_contingency_days
    # h23_1.font = Font(bold=True)
    h23_1.alignment = Alignment(horizontal='center')
    h23_1.border = Border(bottom=hair, left=thin, right=thin)

    ws1.merge_cells('C24:C25')
    top_left_appraisal_cont_1 = ws1['C24']
    top_left_appraisal_cont_1.value = 'Appraisal Contingency'
    # top_left_appraisal_cont_1.font = Font(bold=True)
    top_left_appraisal_cont_1.alignment = Alignment(horizontal='left', vertical='center')
    top_left_appraisal_cont_1.border = Border(top=hair, bottom=hair, left=thin)
    c24_1 = ws1['C24']
    c24_1.border = Border(top=hair, left=thin)
    d24_1 = ws1['D24']
    d24_1.border = Border(top=hair)
    c25_1 = ws1['C25']
    c25_1.border = Border(bottom=hair, left=thin)
    d25_1 = ws1['D25']
    d25_1.border = Border(bottom=hair)

    e24_1 = ws1['E24']
    e24_1.value = offer_1_appraisal_contingency_check
    # e24_1.font = Font(bold=True)
    e24_1.alignment = Alignment(horizontal='center')
    e24_1.border = Border(top=hair, left=thin, right=thin)

    f24_1 = ws1['F24']
    f24_1.value = offer_2_appraisal_contingency_check
    # f24_1.font = Font(bold=True)
    f24_1.alignment = Alignment(horizontal='center')
    f24_1.border = Border(top=hair, left=thin, right=thin)

    g24_1 = ws1['G24']
    g24_1.value = offer_3_appraisal_contingency_check
    # g24_1.font = Font(bold=True)
    g24_1.alignment = Alignment(horizontal='center')
    g24_1.border = Border(top=hair, left=thin, right=thin)

    h24_1 = ws1['H24']
    h24_1.value = offer_4_appraisal_contingency_check
    # h24_1.font = Font(bold=True)
    h24_1.alignment = Alignment(horizontal='center')
    h24_1.border = Border(top=hair, left=thin, right=thin)

    e25_1 = ws1['E25']
    e25_1.value = offer_1_appraisal_contingency_days
    # e25_1.font = Font(bold=True)
    e25_1.alignment = Alignment(horizontal='center')
    e25_1.border = Border(bottom=hair, left=thin, right=thin)

    f25_1 = ws1['F25']
    f25_1.value = offer_2_appraisal_contingency_days
    # f25_1.font = Font(bold=True)
    f25_1.alignment = Alignment(horizontal='center')
    f25_1.border = Border(bottom=hair, left=thin, right=thin)

    g25_1 = ws1['G25']
    g25_1.value = offer_3_appraisal_contingency_days
    # g25_1.font = Font(bold=True)
    g25_1.alignment = Alignment(horizontal='center')
    g25_1.border = Border(bottom=hair, left=thin, right=thin)

    h25_1 = ws1['H25']
    h25_1.value = offer_4_appraisal_contingency_days
    # h25_1.font = Font(bold=True)
    h25_1.alignment = Alignment(horizontal='center')
    h25_1.border = Border(bottom=hair, left=thin, right=thin)

    ws1.merge_cells('C26:C27')
    top_left_home_sale_cont_1 = ws1['C26']
    top_left_home_sale_cont_1.value = 'Home Sale Contingency'
    # top_left_home_sale_cont_1.font = Font(bold=True)
    top_left_home_sale_cont_1.alignment = Alignment(horizontal='left', vertical='center')
    top_left_home_sale_cont_1.border = Border(top=hair, bottom=hair, left=thin)
    c26_1 = ws1['C26']
    c26_1.border = Border(top=hair, left=thin)
    d26_1 = ws1['D26']
    d26_1.border = Border(top=hair)
    c27_1 = ws1['C27']
    c27_1.border = Border(bottom=hair, left=thin)
    d27_1 = ws1['D27']
    d27_1.border = Border(bottom=hair)

    e26_1 = ws1['E26']
    e26_1.value = offer_1_home_sale_contingency_check
    # e26_1.font = Font(bold=True)
    e26_1.alignment = Alignment(horizontal='center')
    e26_1.border = Border(top=hair, left=thin, right=thin)

    f26_1 = ws1['F26']
    f26_1.value = offer_2_home_sale_contingency_check
    # f26_1.font = Font(bold=True)
    f26_1.alignment = Alignment(horizontal='center')
    f26_1.border = Border(top=hair, left=thin, right=thin)

    g26_1 = ws1['G26']
    g26_1.value = offer_3_home_sale_contingency_check
    # g26_1.font = Font(bold=True)
    g26_1.alignment = Alignment(horizontal='center')
    g26_1.border = Border(top=hair, left=thin, right=thin)

    h26_1 = ws1['H26']
    h26_1.value = offer_4_home_sale_contingency_check
    # h26_1.font = Font(bold=True)
    h26_1.alignment = Alignment(horizontal='center')
    h26_1.border = Border(top=hair, left=thin, right=thin)

    e27_1 = ws1['E27']
    e27_1.value = offer_1_home_sale_contingency_days
    # e27_1.font = Font(bold=True)
    e27_1.alignment = Alignment(horizontal='center')
    e27_1.border = Border(bottom=hair, left=thin, right=thin)

    f27_1 = ws1['F27']
    f27_1.value = offer_2_home_sale_contingency_days
    # f27_1.font = Font(bold=True)
    f27_1.alignment = Alignment(horizontal='center')
    f27_1.border = Border(bottom=hair, left=thin, right=thin)

    g27_1 = ws1['G27']
    g27_1.value = offer_3_home_sale_contingency_days
    # g27_1.font = Font(bold=True)
    g27_1.alignment = Alignment(horizontal='center')
    g27_1.border = Border(bottom=hair, left=thin, right=thin)

    h27_1 = ws1['H27']
    h27_1.value = offer_4_home_sale_contingency_days
    # h27_1.font = Font(bold=True)
    h27_1.alignment = Alignment(horizontal='center')
    h27_1.border = Border(bottom=hair, left=thin, right=thin)

    c28_1 = ws1['C28']
    c28_1.value = 'Pre Occupancy Start Date'
    # c28_1.font = Font(bold=True)
    c28_1.alignment = Alignment(horizontal='left', vertical='center')
    c28_1.border = Border(top=hair, bottom=hair, left=thin)

    d28_1 = ws1['D28']
    d28_1.border = Border(top=hair, bottom=hair)

    e28_1 = ws1['E28']
    e28_1.value = offer_1_pre_occupancy_date
    # e28_1.font = Font(bold=True)
    e28_1.alignment = Alignment(horizontal='right')
    e28_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f28_1 = ws1['F28']
    f28_1.value = offer_2_pre_occupancy_date
    # f28_1.font = Font(bold=True)
    f28_1.alignment = Alignment(horizontal='right')
    f28_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g28_1 = ws1['G28']
    g28_1.value = offer_3_pre_occupancy_date
    # g28_1.font = Font(bold=True)
    g28_1.alignment = Alignment(horizontal='right')
    g28_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h28_1 = ws1['H28']
    h28_1.value = offer_4_pre_occupancy_date
    # h28_1.font = Font(bold=True)
    h28_1.alignment = Alignment(horizontal='right')
    h28_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c29_1 = ws1['C29']
    c29_1.value = 'Post Occupancy Thru Date'
    # c29_1.font = Font(bold=True)
    c29_1.alignment = Alignment(horizontal='left', vertical='center')
    c29_1.border = Border(top=hair, bottom=thin, left=thin)

    d29_1 = ws1['D29']
    d29_1.border = Border(top=hair, bottom=thin)

    e29_1 = ws1['E29']
    e29_1.value = offer_1_post_occupancy_date
    # e29_1.font = Font(bold=True)
    e29_1.alignment = Alignment(horizontal='right')
    e29_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f29_1 = ws1['F29']
    f29_1.value = offer_2_post_occupancy_date
    # f29_1.font = Font(bold=True)
    f29_1.alignment = Alignment(horizontal='right')
    f29_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g29_1 = ws1['G29']
    g29_1.value = offer_3_post_occupancy_date
    # g29_1.font = Font(bold=True)
    g29_1.alignment = Alignment(horizontal='right')
    g29_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h29_1 = ws1['H29']
    h29_1.value = offer_4_post_occupancy_date
    # h29_1.font = Font(bold=True)
    h29_1.alignment = Alignment(horizontal='right')
    h29_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    # Housing Related Cost Table
    c31_1 = ws1['C31']
    c31_1.value = 'HOUSING-RELATED COSTS'
    c31_1.font = Font(bold=True)
    c31_1.border = Border(bottom=thin)

    d31_1 = ws1['D31']
    d31_1.value = 'Calculation Description'
    d31_1.font = Font(bold=True)
    d31_1.border = Border(bottom=thin)

    c32_1 = ws1['C32']
    c32_1.value = 'Estimated Payoff - 1st Trust'
    c32_1.border = Border(top=thin, bottom=hair, left=thin)

    d32_1 = ws1['D32']
    d32_1.value = 'Principal Balance of Loan'
    d32_1.border = Border(top=thin, bottom=hair)

    e32_1 = ws1['E32']
    e32_1.value = first_trust
    e32_1.number_format = acct_fmt
    e32_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f32_1 = ws1['F32']
    f32_1.value = first_trust
    f32_1.number_format = acct_fmt
    f32_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g32_1 = ws1['G32']
    g32_1.value = first_trust
    g32_1.number_format = acct_fmt
    g32_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h32_1 = ws1['H32']
    h32_1.value = first_trust
    h32_1.number_format = acct_fmt
    h32_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c33_1 = ws1['C33']
    c33_1.value = 'Estimated Payoff - 2nd Trust'
    c33_1.border = Border(top=hair, bottom=hair, left=thin)

    d33_1 = ws1['D33']
    d33_1.value = 'Principal Balance of Loan'
    d33_1.border = Border(top=hair, bottom=hair)

    e33_1 = ws1['E33']
    e33_1.value = second_trust
    e33_1.number_format = acct_fmt
    e33_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f33_1 = ws1['F33']
    f33_1.value = second_trust
    f33_1.number_format = acct_fmt
    f33_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g33_1 = ws1['G33']
    g33_1.value = second_trust
    g33_1.number_format = acct_fmt
    g33_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h33_1 = ws1['H33']
    h33_1.value = second_trust
    h33_1.number_format = acct_fmt
    h33_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c34_1 = ws1['C34']
    c34_1.value = 'Purchaser Closing Cost / Contract'
    c34_1.border = Border(top=hair, bottom=hair, left=thin)

    d34_1 = ws1['D34']
    d34_1.value = 'Negotiated Into Contract'
    d34_1.border = Border(top=hair, bottom=hair)

    e34_1 = ws1['E34']
    e34_1.value = offer_1_closing_cost_subsidy_amt
    e34_1.number_format = acct_fmt
    e34_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f34_1 = ws1['F34']
    f34_1.value = offer_2_closing_cost_subsidy_amt
    f34_1.number_format = acct_fmt
    f34_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g34_1 = ws1['G34']
    g34_1.value = offer_3_closing_cost_subsidy_amt
    g34_1.number_format = acct_fmt
    g34_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h34_1 = ws1['H34']
    h34_1.value = offer_4_closing_cost_subsidy_amt
    h34_1.number_format = acct_fmt
    h34_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c35_1 = ws1['C35']
    c35_1.value = 'Prorated Taxes / Assessments'
    c35_1.border = Border(top=hair, bottom=hair, left=thin)

    d35_1 = ws1['D35']
    d35_1.value = '1 Year of Taxes Divided by 12 Multiplied by 3'
    d35_1.border = Border(top=hair, bottom=hair)

    e35_1 = ws1['E35']
    e35_1.value = prorated_taxes
    e35_1.number_format = acct_fmt
    e35_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f35_1 = ws1['F35']
    f35_1.value = prorated_taxes
    f35_1.number_format = acct_fmt
    f35_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g35_1 = ws1['G35']
    g35_1.value = prorated_taxes
    g35_1.number_format = acct_fmt
    g35_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h35_1 = ws1['H35']
    h35_1.value = prorated_taxes
    h35_1.number_format = acct_fmt
    h35_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c36_1 = ws1['C36']
    c36_1.value = 'Prorated HOA / Condo Dues'
    c36_1.border = Border(top=hair, bottom=thin, left=thin)

    d36_1 = ws1['D36']
    d36_1.value = '1 Year of Dues Divided by 12 Multiplied by 3'
    d36_1.border = Border(top=hair, bottom=thin)

    e36_1 = ws1['E36']
    e36_1.value = prorated_hoa_condo_fees
    e36_1.number_format = acct_fmt
    e36_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f36_1 = ws1['F36']
    f36_1.value = prorated_hoa_condo_fees
    f36_1.number_format = acct_fmt
    f36_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g36_1 = ws1['G36']
    g36_1.value = prorated_hoa_condo_fees
    g36_1.number_format = acct_fmt
    g36_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h36_1 = ws1['H36']
    h36_1.value = prorated_hoa_condo_fees
    h36_1.number_format = acct_fmt
    h36_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e37_1 = ws1['E37']
    e37_1.value = '=SUM(E32:E36)'
    e37_1.font = Font(bold=True)
    e37_1.number_format = acct_fmt
    e37_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f37_1 = ws1['F37']
    f37_1.value = '=SUM(F32:F36)'
    f37_1.font = Font(bold=True)
    f37_1.number_format = acct_fmt
    f37_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g37_1 = ws1['G37']
    g37_1.value = '=SUM(G32:G36)'
    g37_1.font = Font(bold=True)
    g37_1.number_format = acct_fmt
    g37_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h37_1 = ws1['H37']
    h37_1.value = '=SUM(H32:H36)'
    h37_1.font = Font(bold=True)
    h37_1.number_format = acct_fmt
    h37_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Broker and Financing Costs Table
    c38_1 = ws1['C38']
    c38_1.value = 'BROKERAGE & FINANCING COSTS'
    c38_1.font = Font(bold=True)
    c38_1.border = Border(bottom=thin)

    c39_1 = ws1['C39']
    c39_1.value = 'Listing Company Compensation'
    c39_1.border = Border(top=thin, bottom=hair, left=thin)

    d39_1 = ws1['D39']
    d39_1.value = '% from Listing Agreement * Offer Amount ($)'
    d39_1.border = Border(top=thin, bottom=hair)

    e39_1 = ws1['E39']
    e39_1.value = listing_company_pct * offer_1_amt
    e39_1.number_format = acct_fmt
    e39_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f39_1 = ws1['F39']
    f39_1.value = listing_company_pct * offer_2_amt
    f39_1.number_format = acct_fmt
    f39_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g39_1 = ws1['G39']
    g39_1.value = listing_company_pct * offer_3_amt
    g39_1.number_format = acct_fmt
    g39_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h39_1 = ws1['H39']
    h39_1.value = listing_company_pct * offer_4_amt
    h39_1.number_format = acct_fmt
    h39_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c40_1 = ws1['C40']
    c40_1.value = 'Selling Company Compensation'
    c40_1.border = Border(top=hair, bottom=hair, left=thin)

    d40_1 = ws1['D40']
    d40_1.value = '% from Listing Agreement * Offer Amount ($)'
    d40_1.border = Border(top=hair, bottom=hair)

    e40_1 = ws1['E40']
    e40_1.value = selling_company_pct * offer_1_amt
    e40_1.number_format = acct_fmt
    e40_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f40_1 = ws1['F40']
    f40_1.value = selling_company_pct * offer_2_amt
    f40_1.number_format = acct_fmt
    f40_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g40_1 = ws1['G40']
    g40_1.value = selling_company_pct * offer_3_amt
    g40_1.number_format = acct_fmt
    g40_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h40_1 = ws1['H40']
    h40_1.value = selling_company_pct * offer_4_amt
    h40_1.number_format = acct_fmt
    h40_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c41_1 = ws1['C41']
    c41_1.value = 'Processing Fee'
    c41_1.border = Border(top=hair, bottom=thin, left=thin)

    d41_1 = ws1['D41']
    d41_1.border = Border(top=hair, bottom=thin)

    e41_1 = ws1['E41']
    e41_1.value = processing_fee
    e41_1.number_format = acct_fmt
    e41_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f41_1 = ws1['F41']
    f41_1.value = processing_fee
    f41_1.number_format = acct_fmt
    f41_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g41_1 = ws1['G41']
    g41_1.value = processing_fee
    g41_1.number_format = acct_fmt
    g41_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h41_1 = ws1['H41']
    h41_1.value = processing_fee
    h41_1.number_format = acct_fmt
    h41_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e42_1 = ws1['E42']
    e42_1.value = '=SUM(E39:E41)'
    e42_1.font = Font(bold=True)
    e42_1.number_format = acct_fmt
    e42_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f42_1 = ws1['F42']
    f42_1.value = '=SUM(F39:F41)'
    f42_1.font = Font(bold=True)
    f42_1.number_format = acct_fmt
    f42_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g42_1 = ws1['G42']
    g42_1.value = '=SUM(G39:G41)'
    g42_1.font = Font(bold=True)
    g42_1.number_format = acct_fmt
    g42_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h42_1 = ws1['H42']
    h42_1.value = '=SUM(H39:H41)'
    h42_1.font = Font(bold=True)
    h42_1.number_format = acct_fmt
    h42_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Estimated Closing Costs Table
    c43_1 = ws1['C43']
    c43_1.value = 'ESTIMATED CLOSING COSTS'
    c43_1.font = Font(bold=True)
    c43_1.border = Border(bottom=thin)

    c44_1 = ws1['C44']
    c44_1.value = 'Settlement Fee'
    c44_1.border = Border(top=thin, bottom=hair, left=thin)

    d44_1 = ws1['D44']
    d44_1.value = 'Commonly Used Fee'
    d44_1.border = Border(top=thin, bottom=hair)

    e44_1 = ws1['E44']
    e44_1.value = settlement_fee
    e44_1.number_format = acct_fmt
    e44_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f44_1 = ws1['F44']
    f44_1.value = settlement_fee
    f44_1.number_format = acct_fmt
    f44_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g44_1 = ws1['G44']
    g44_1.value = settlement_fee
    g44_1.number_format = acct_fmt
    g44_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h44_1 = ws1['H44']
    h44_1.value = settlement_fee
    h44_1.number_format = acct_fmt
    h44_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c45_1 = ws1['C45']
    c45_1.value = 'Deed Preparation'
    c45_1.border = Border(top=hair, bottom=hair, left=thin)

    d45_1 = ws1['D45']
    d45_1.value = 'Commonly Used Fee'
    d45_1.border = Border(top=hair, bottom=hair)

    e45_1 = ws1['E45']
    e45_1.value = deed_preparation_fee
    e45_1.number_format = acct_fmt
    e45_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f45_1 = ws1['F45']
    f45_1.value = deed_preparation_fee
    f45_1.number_format = acct_fmt
    f45_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g45_1 = ws1['G45']
    g45_1.value = deed_preparation_fee
    g45_1.number_format = acct_fmt
    g45_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h45_1 = ws1['H45']
    h45_1.value = deed_preparation_fee
    h45_1.number_format = acct_fmt
    h45_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c46_1 = ws1['C46']
    c46_1.value = 'Release of Liens / Trusts'
    c46_1.border = Border(top=hair, bottom=thin, left=thin)

    d46_1 = ws1['D46']
    d46_1.value = 'Commonly Used Fee * Qty of Trusts or Liens'
    d46_1.border = Border(top=hair, bottom=thin)

    e46_1 = ws1['E46']
    e46_1.value = lien_trust_release_fee * lien_trust_release_qty
    e46_1.number_format = acct_fmt
    e46_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f46_1 = ws1['F46']
    f46_1.value = lien_trust_release_fee * lien_trust_release_qty
    f46_1.number_format = acct_fmt
    f46_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g46_1 = ws1['G46']
    g46_1.value = lien_trust_release_fee * lien_trust_release_qty
    g46_1.number_format = acct_fmt
    g46_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h46_1 = ws1['H46']
    h46_1.value = lien_trust_release_fee * lien_trust_release_qty
    h46_1.number_format = acct_fmt
    h46_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e47_1 = ws1['E47']
    e47_1.value = '=SUM(E44:E46)'
    e47_1.font = Font(bold=True)
    e47_1.number_format = acct_fmt
    e47_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f47_1 = ws1['F47']
    f47_1.value = '=SUM(F44:F46)'
    f47_1.font = Font(bold=True)
    f47_1.number_format = acct_fmt
    f47_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g47_1 = ws1['G47']
    g47_1.value = '=SUM(G44:G46)'
    g47_1.font = Font(bold=True)
    g47_1.number_format = acct_fmt
    g47_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h47_1 = ws1['H47']
    h47_1.value = '=SUM(H44:H46)'
    h47_1.font = Font(bold=True)
    h47_1.number_format = acct_fmt
    h47_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Misc Costs Table
    c48_1 = ws1['C48']
    c48_1.value = 'MISCELLANEOUS COSTS'
    c48_1.font = Font(bold=True)
    c48_1.border = Border(bottom=thin)

    c49_1 = ws1['C49']
    c49_1.value = 'Recording Release(s)'
    c49_1.border = Border(top=thin, bottom=hair, left=thin)

    d49_1 = ws1['D49']
    d49_1.value = 'Commonly Used Fee * Qty of Trusts Recorded'
    d49_1.border = Border(top=thin, bottom=hair)

    e49_1 = ws1['E49']
    e49_1.value = recording_fee * recording_trusts_liens_qty
    e49_1.number_format = acct_fmt
    e49_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f49_1 = ws1['F49']
    f49_1.value = recording_fee * recording_trusts_liens_qty
    f49_1.number_format = acct_fmt
    f49_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g49_1 = ws1['G49']
    g49_1.value = recording_fee * recording_trusts_liens_qty
    g49_1.number_format = acct_fmt
    g49_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h49_1 = ws1['H49']
    h49_1.value = recording_fee * recording_trusts_liens_qty
    h49_1.number_format = acct_fmt
    h49_1.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c50_1 = ws1['C50']
    c50_1.value = 'Grantor\'s Tax'
    c50_1.border = Border(top=hair, bottom=hair, left=thin)

    d50_1 = ws1['D50']
    d50_1.value = '% of Offer Amount ($)'
    d50_1.border = Border(top=hair, bottom=hair)

    e50_1 = ws1['E50']
    e50_1.value = grantors_tax_pct * offer_1_amt
    e50_1.number_format = acct_fmt
    e50_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f50_1 = ws1['F50']
    f50_1.value = grantors_tax_pct * offer_2_amt
    f50_1.number_format = acct_fmt
    f50_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g50_1 = ws1['G50']
    g50_1.value = grantors_tax_pct * offer_3_amt
    g50_1.number_format = acct_fmt
    g50_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h50_1 = ws1['H50']
    h50_1.value = grantors_tax_pct * offer_4_amt
    h50_1.number_format = acct_fmt
    h50_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c51_1 = ws1['C51']
    c51_1.value = 'Congestion Relief Tax'
    c51_1.border = Border(top=hair, bottom=hair, left=thin)

    d51_1 = ws1['D51']
    d51_1.value = '% of Offer Amount ($)'
    d51_1.border = Border(top=hair, bottom=hair)

    e51_1 = ws1['E51']
    e51_1.value = congestion_tax_pct * offer_1_amt
    e51_1.number_format = acct_fmt
    e51_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f51_1 = ws1['F51']
    f51_1.value = congestion_tax_pct * offer_2_amt
    f51_1.number_format = acct_fmt
    f51_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g51_1 = ws1['G51']
    g51_1.value = congestion_tax_pct * offer_3_amt
    g51_1.number_format = acct_fmt
    g51_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h51_1 = ws1['H51']
    h51_1.value = congestion_tax_pct * offer_4_amt
    h51_1.number_format = acct_fmt
    h51_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c52_1 = ws1['C52']
    c52_1.value = 'Pest Inspection'
    c52_1.border = Border(top=hair, bottom=hair, left=thin)

    d52_1 = ws1['D52']
    d52_1.value = 'Commonly Used Fee'
    d52_1.border = Border(top=hair, bottom=hair)

    e52_1 = ws1['E52']
    e52_1.value = pest_inspection_fee
    e52_1.number_format = acct_fmt
    e52_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f52_1 = ws1['F52']
    f52_1.value = pest_inspection_fee
    f52_1.number_format = acct_fmt
    f52_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g52_1 = ws1['G52']
    g52_1.value = pest_inspection_fee
    g52_1.number_format = acct_fmt
    g52_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h52_1 = ws1['H52']
    h52_1.value = pest_inspection_fee
    h52_1.number_format = acct_fmt
    h52_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c53_1 = ws1['C53']
    c53_1.value = 'POA / Condo Disclosures'
    c53_1.border = Border(top=hair, bottom=hair, left=thin)

    d53_1 = ws1['D53']
    d53_1.value = 'Commonly Used Fee'
    d53_1.border = Border(top=hair, bottom=hair)

    e53_1 = ws1['E53']
    e53_1.value = poa_condo_disclosure_fee
    e53_1.number_format = acct_fmt
    e53_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f53_1 = ws1['F53']
    f53_1.value = poa_condo_disclosure_fee
    f53_1.number_format = acct_fmt
    f53_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g53_1 = ws1['G53']
    g53_1.value = poa_condo_disclosure_fee
    g53_1.number_format = acct_fmt
    g53_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h53_1 = ws1['H53']
    h53_1.value = poa_condo_disclosure_fee
    h53_1.number_format = acct_fmt
    h53_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c54_1 = ws1['C54']
    c54_1.value = 'Pre Occupancy Credit to Seller'
    c54_1.border = Border(top=hair, bottom=hair, left=thin)

    d54_1 = ws1['D54']
    d54_1.value = 'Negotiated Into Contract'
    d54_1.border = Border(top=hair, bottom=hair)

    e54_1 = ws1['E54']
    e54_1.value = offer_1_pre_occupancy_credit_amt
    e54_1.number_format = acct_fmt
    e54_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f54_1 = ws1['F54']
    f54_1.value = offer_2_pre_occupancy_credit_amt
    f54_1.number_format = acct_fmt
    f54_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g54_1 = ws1['G54']
    g54_1.value = offer_3_pre_occupancy_credit_amt
    g54_1.number_format = acct_fmt
    g54_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h54_1 = ws1['H54']
    h54_1.value = offer_4_pre_occupancy_credit_amt
    h54_1.number_format = acct_fmt
    h54_1.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c55_1 = ws1['C55']
    c55_1.value = 'Post Occupancy Cost to Seller'
    c55_1.border = Border(top=hair, bottom=thin, left=thin)

    d55_1 = ws1['D55']
    d55_1.value = 'Negotiated Into Contract'
    d55_1.border = Border(top=hair, bottom=thin)

    e55_1 = ws1['E55']
    e55_1.value = offer_1_post_occupancy_cost_amt
    e55_1.number_format = acct_fmt
    e55_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f55_1 = ws1['F55']
    f55_1.value = offer_2_post_occupancy_cost_amt
    f55_1.number_format = acct_fmt
    f55_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g55_1 = ws1['G55']
    g55_1.value = offer_3_post_occupancy_cost_amt
    g55_1.number_format = acct_fmt
    g55_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h55_1 = ws1['H55']
    h55_1.value = offer_4_post_occupancy_cost_amt
    h55_1.number_format = acct_fmt
    h55_1.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e56_1 = ws1['E56']
    e56_1.value = '=SUM(E49:E53,E55)-E54'
    e56_1.font = Font(bold=True)
    e56_1.number_format = acct_fmt
    e56_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f56_1 = ws1['F56']
    f56_1.value = '=SUM(F49:F53,F55)-F54'
    f56_1.font = Font(bold=True)
    f56_1.number_format = acct_fmt
    f56_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g56_1 = ws1['G56']
    g56_1.value = '=SUM(G49:G53,G55)-G54'
    g56_1.font = Font(bold=True)
    g56_1.number_format = acct_fmt
    g56_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h56_1 = ws1['H56']
    h56_1.value = '=SUM(H49:H53,H55)-H54'
    h56_1.font = Font(bold=True)
    h56_1.number_format = acct_fmt
    h56_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Subtotals section
    ws1.merge_cells('C58:D58')
    top_left_cell_four_1 = ws1['C58']
    top_left_cell_four_1.value = 'TOTAL ESTIMATED COST OF SETTLEMENT'
    top_left_cell_four_1.font = Font(bold=True)
    top_left_cell_four_1.alignment = Alignment(horizontal='right')

    e58_1 = ws1['E58']
    e58_1.value = '=SUM(E37,E42,E47,E56)'
    e58_1.font = Font(bold=True)
    e58_1.number_format = acct_fmt
    e58_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f58_1 = ws1['F58']
    f58_1.value = '=SUM(F37,F42,F47,F56)'
    f58_1.font = Font(bold=True)
    f58_1.number_format = acct_fmt
    f58_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g58_1 = ws1['G58']
    g58_1.value = '=SUM(G37,G42,G47,G56)'
    g58_1.font = Font(bold=True)
    g58_1.number_format = acct_fmt
    g58_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h58_1 = ws1['H58']
    h58_1.value = '=SUM(H37,H42,H47,H56)'
    h58_1.font = Font(bold=True)
    h58_1.number_format = acct_fmt
    h58_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws1.merge_cells('C59:D59')
    top_left_cell_five_1 = ws1['C59']
    top_left_cell_five_1.value = 'Offer Amount ($)'
    top_left_cell_five_1.font = Font(bold=True)
    top_left_cell_five_1.alignment = Alignment(horizontal='right')

    e59_1 = ws1['E59']
    e59_1.value = offer_1_amt
    e59_1.font = Font(bold=True)
    e59_1.number_format = acct_fmt
    e59_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f59_1 = ws1['F59']
    f59_1.value = offer_2_amt
    f59_1.font = Font(bold=True)
    f59_1.number_format = acct_fmt
    f59_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g59_1 = ws1['G59']
    g59_1.value = offer_3_amt
    g59_1.font = Font(bold=True)
    g59_1.number_format = acct_fmt
    g59_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h59_1 = ws1['H59']
    h59_1.value = offer_4_amt
    h59_1.font = Font(bold=True)
    h59_1.number_format = acct_fmt
    h59_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws1.merge_cells('C60:D60')
    top_left_cell_six_1 = ws1['C60']
    top_left_cell_six_1.value = 'LESS: Total Estimated Cost of Settlement'
    top_left_cell_six_1.font = Font(bold=True)
    top_left_cell_six_1.alignment = Alignment(horizontal='right')

    e60_1 = ws1['E60']
    e60_1.value = '=-SUM(E37,E42,E47,E56)'
    e60_1.font = Font(bold=True)
    e60_1.number_format = acct_fmt
    e60_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f60_1 = ws1['F60']
    f60_1.value = '=-SUM(F37,F42,F47,F56)'
    f60_1.font = Font(bold=True)
    f60_1.number_format = acct_fmt
    f60_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g60_1 = ws1['G60']
    g60_1.value = '=-SUM(G37,G42,G47,G56)'
    g60_1.font = Font(bold=True)
    g60_1.number_format = acct_fmt
    g60_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h60_1 = ws1['H60']
    h60_1.value = '=-SUM(H37,H42,H47,H56)'
    h60_1.font = Font(bold=True)
    h60_1.number_format = acct_fmt
    h60_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws1.merge_cells('C61:D61')
    top_left_cell_seven_1 = ws1['C61']
    top_left_cell_seven_1.value = 'ESTIMATED TOTAL NET PROCEEDS'
    top_left_cell_seven_1.font = Font(bold=True)
    top_left_cell_seven_1.alignment = Alignment(horizontal='right')

    e61_1 = ws1['E61']
    e61_1.value = '=SUM(E59:E60)'
    e61_1.font = Font(bold=True)
    e61_1.number_format = acct_fmt
    e61_1.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    e61_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f61_1 = ws1['F61']
    f61_1.value = '=SUM(F59:F60)'
    f61_1.font = Font(bold=True)
    f61_1.number_format = acct_fmt
    f61_1.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    f61_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g61_1 = ws1['G61']
    g61_1.value = '=SUM(G59:G60)'
    g61_1.font = Font(bold=True)
    g61_1.number_format = acct_fmt
    g61_1.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    g61_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h61_1 = ws1['H61']
    h61_1.value = '=SUM(H59:H60)'
    h61_1.font = Font(bold=True)
    h61_1.number_format = acct_fmt
    h61_1.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    h61_1.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Signature Block
    c63_1 = ws1['C63']
    c63_1.value = 'PREPARED BY:'

    c64_1 = ws1['C64']
    c64_1.value = agent

    e63_1 = ws1['E63']
    e63_1.value = 'SELLER:'

    e64_1 = ws1['E64']
    e64_1.value = seller_name

    # Freedom Logo
    # c53 = ws1['C53']
    freedom_logo = drawing.image.Image('freedom_logo.png')
    freedom_logo.height = 80
    freedom_logo.width = 115
    freedom_logo.anchor = 'C66'
    ws1.add_image(freedom_logo)

    # Disclosure Statement
    ws1.merge_cells('C70:H74')
    top_left_cell_eight_1 = ws1['C70']
    top_left_cell_eight_1.value = '''These estimates are not guaranteed and may not include escrows. Escrow balances are reimbursed by the existing lender. Taxes, rents & association dues are pro-rated at settlement. Under Virginia Law, the seller's proceeds may not be available for up to 2 business days following recording of the deed. Seller acknowledges receipt of this statement.'''
    top_left_cell_eight_1.font = Font(italic=True)
    top_left_cell_eight_1.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

    # page 2
    # Build Header
    ws2.merge_cells('C2:H2')
    top_left_cell_one_2 = ws2['C2']
    top_left_cell_one_2.value = 'Seller\'s Total Net Proceeds For Different Offers'
    top_left_cell_one_2.font = Font(bold=True)
    top_left_cell_one_2.alignment = Alignment(horizontal='center')

    ws2.merge_cells('C3:H3')
    top_left_cell_two_2 = ws2['C3']
    top_left_cell_two_2.value = f'{seller_name} - {seller_address}'
    top_left_cell_two_2.font = Font(bold=True)
    top_left_cell_two_2.alignment = Alignment(horizontal='center')

    ws2.merge_cells('C4:H4')
    top_left_cell_three_2 = ws2['C4']
    top_left_cell_three_2.value = f'Date Prepared: {date}'
    top_left_cell_three_2.font = Font(bold=True)
    top_left_cell_three_2.alignment = Alignment(horizontal='center')

    ws2.merge_cells('C5:H5')
    top_left_cell_four_2 = ws2['C5']
    top_left_cell_four_2.value = f'List Price: ${list_price:,.2f}'
    top_left_cell_four_2.font = Font(bold=True)
    top_left_cell_four_2.alignment = Alignment(horizontal='center')

    c7_2 = ws2['C7']
    c7_2.value = 'OFFER SUMMARY FEATURES'
    c7_2.font = Font(bold=True)
    c7_2.border = Border(bottom=thin)

    e7_2 = ws2['E7']
    e7_2.value = offer_2_name
    e7_2.font = Font(bold=True)
    e7_2.border = Border(bottom=thin)
    e7_2.alignment = Alignment(horizontal='center', wrap_text=True)

    f7_2 = ws2['F7']
    f7_2.value = offer_2_name
    f7_2.font = Font(bold=True)
    f7_2.border = Border(bottom=thin)
    f7_2.alignment = Alignment(horizontal='center', wrap_text=True)

    g7_2 = ws2['G7']
    g7_2.value = offer_3_name
    g7_2.font = Font(bold=True)
    g7_2.border = Border(bottom=thin)
    g7_2.alignment = Alignment(horizontal='center', wrap_text=True)

    h7_2 = ws2['H7']
    h7_2.value = offer_4_name
    h7_2.font = Font(bold=True)
    h7_2.border = Border(bottom=thin)
    h7_2.alignment = Alignment(horizontal='center', wrap_text=True)

    c8_2 = ws2['C8']
    c8_2.value = 'Offer Amt. ($)'
    # c8_2.font = Font(bold=True)
    c8_2.alignment = Alignment(horizontal='left', vertical='center')
    c8_2.border = Border(top=thin, bottom=hair, left=thin)

    d8_2 = ws2['D8']
    d8_2.border = Border(top=thin, bottom=hair)

    e8_2 = ws2['E8']
    e8_2.value = offer_2_amt
    # e8_2.font = Font(bold=True)
    e8_2.number_format = acct_fmt
    e8_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f8_2 = ws2['F8']
    f8_2.value = offer_2_amt
    # f8_2.font = Font(bold=True)
    f8_2.number_format = acct_fmt
    f8_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g8_2 = ws2['G8']
    g8_2.value = offer_3_amt
    # g8_2.font = Font(bold=True)
    g8_2.number_format = acct_fmt
    g8_2.border = Border(top=thin, bottom=hair, right=thin, left=thin)

    h8_2 = ws2['H8']
    h8_2.value = offer_4_amt
    # h8_2.font = Font(bold=True)
    h8_2.number_format = acct_fmt
    h8_2.border = Border(top=thin, bottom=hair, right=thin, left=thin)

    c9_2 = ws2['C9']
    c9_2.value = 'Down Pmt (%)'
    # c9_2.font = Font(bold=True)
    c9_2.alignment = Alignment(horizontal='left', vertical='center')
    c9_2.border = Border(top=hair, bottom=hair, left=thin)

    d9_2 = ws2['D9']
    d9_2.border = Border(top=hair, bottom=hair)

    e9_2 = ws2['E9']
    e9_2.value = offer_2_down_pmt_pct
    # e9_2.font = Font(bold=True)
    e9_2.number_format = pct_fmt
    e9_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f9_2 = ws2['F9']
    f9_2.value = offer_2_down_pmt_pct
    # f9_2.font = Font(bold=True)
    f9_2.number_format = pct_fmt
    f9_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g9_2 = ws2['G9']
    g9_2.value = offer_3_down_pmt_pct
    # g9_2.font = Font(bold=True)
    g9_2.number_format = pct_fmt
    g9_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h9_2 = ws2['H9']
    h9_2.value = offer_4_down_pmt_pct
    # h9_2.font = Font(bold=True)
    h9_2.number_format = pct_fmt
    h9_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c10_2 = ws2['C10']
    c10_2.value = 'Settlement Date'
    # c10_2.font = Font(bold=True)
    c10_2.alignment = Alignment(horizontal='left', vertical='center')
    c10_2.border = Border(top=hair, bottom=hair, left=thin)

    d10_2 = ws2['D10']
    d10_2.border = Border(top=hair, bottom=hair)

    e10_2 = ws2['E10']
    e10_2.value = offer_2_settlement_date
    # e10_2.font = Font(bold=True)
    e10_2.alignment = Alignment(horizontal='right')
    e10_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f10_2 = ws2['F10']
    f10_2.value = offer_2_settlement_date
    # f10_2.font = Font(bold=True)
    f10_2.alignment = Alignment(horizontal='right')
    f10_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g10_2 = ws2['G10']
    g10_2.value = offer_3_settlement_date
    # g10_2.font = Font(bold=True)
    g10_2.alignment = Alignment(horizontal='right')
    g10_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h10_2 = ws2['H10']
    h10_2.value = offer_4_settlement_date
    # h10_2.font = Font(bold=True)
    h10_2.alignment = Alignment(horizontal='right')
    h10_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c11_2 = ws2['C11']
    c11_2.value = 'Settlement Company'
    # c11_2.font = Font(bold=True)
    c11_2.alignment = Alignment(horizontal='left', vertical='center')
    c11_2.border = Border(top=hair, bottom=hair, left=thin)

    d11_2 = ws2['D11']
    d11_2.border = Border(top=hair, bottom=hair)

    e11_2 = ws2['E11']
    e11_2.value = offer_2_settlement_company
    # e11_2.font = Font(bold=True)
    e11_2.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    e11_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f11_2 = ws2['F11']
    f11_2.value = offer_2_settlement_company
    # f11_2.font = Font(bold=True)
    f11_2.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    f11_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g11_2 = ws2['G11']
    g11_2.value = offer_3_settlement_company
    # g11_2.font = Font(bold=True)
    g11_2.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    g11_2.border = Border(top=hair, bottom=hair, right=thin, left=thin)

    h11_2 = ws2['H11']
    h11_2.value = offer_4_settlement_company
    # h11_2.font = Font(bold=True)
    h11_2.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    h11_2.border = Border(top=hair, bottom=hair, right=thin, left=thin)

    c12_2 = ws2['C12']
    c12_2.value = 'EMD Amt. ($)'
    # c12_2.font = Font(bold=True)
    c12_2.alignment = Alignment(horizontal='left', vertical='center')
    c12_2.border = Border(top=hair, bottom=hair, left=thin)

    d12_2 = ws2['D12']
    d12_2.border = Border(top=hair, bottom=hair)

    e12_2 = ws2['E12']
    e12_2.value = offer_2_emd_amt
    # e12_2.font = Font(bold=True)
    e12_2.number_format = acct_fmt
    e12_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f12_2 = ws2['F12']
    f12_2.value = offer_2_emd_amt
    # f12_2.font = Font(bold=True)
    f12_2.number_format = acct_fmt
    f12_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g12_2 = ws2['G12']
    g12_2.value = offer_3_emd_amt
    # g12_2.font = Font(bold=True)
    g12_2.number_format = acct_fmt
    g12_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h12_2 = ws2['H12']
    h12_2.value = offer_4_emd_amt
    # h12_2.font = Font(bold=True)
    h12_2.number_format = acct_fmt
    h12_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c13_2 = ws2['C13']
    c13_2.value = 'Financing Type'
    # c13_2.font = Font(bold=True)
    c13_2.alignment = Alignment(horizontal='left', vertical='center')
    c13_2.border = Border(top=hair, bottom=hair, left=thin)

    d13_2 = ws2['D13']
    d13_2.border = Border(top=hair, bottom=hair)

    e13_2 = ws2['E13']
    e13_2.value = offer_2_financing_type
    # e13_2.font = Font(bold=True)
    e13_2.alignment = Alignment(horizontal='center')
    e13_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f13_2 = ws2['F13']
    f13_2.value = offer_2_financing_type
    # f13_2.font = Font(bold=True)
    f13_2.alignment = Alignment(horizontal='center')
    f13_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g13_2 = ws2['G13']
    g13_2.value = offer_3_financing_type
    # g13_2.font = Font(bold=True)
    g13_2.alignment = Alignment(horizontal='center')
    g13_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h13_2 = ws2['H13']
    h13_2.value = offer_4_financing_type
    # h13_2.font = Font(bold=True)
    h13_2.alignment = Alignment(horizontal='center')
    h13_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    ws2.merge_cells('C14:C15')
    top_left_home_insp_cont_2 = ws2['C14']
    top_left_home_insp_cont_2.value = 'Home Inspection Contingency'
    # top_left_home_insp_cont_2.font = Font(bold=True)
    top_left_home_insp_cont_2.alignment = Alignment(horizontal='left', vertical='center')
    top_left_home_insp_cont_2.border = Border(top=hair, bottom=hair, left=thin)
    c14_2 = ws2['C14']
    c14_2.border = Border(top=hair, left=thin)
    d14_2 = ws2['D14']
    d14_2.border = Border(top=hair)
    c15_2 = ws2['C15']
    c15_2.border = Border(bottom=hair, left=thin)
    d15_2 = ws2['D15']
    d15_2.border = Border(bottom=hair)

    e14_2 = ws2['E14']
    e14_2.value = offer_2_home_inspection_check
    # e14_2.font = Font(bold=True)
    e14_2.alignment = Alignment(horizontal='center')
    e14_2.border = Border(top=hair, left=thin, right=thin)

    f14_2 = ws2['F14']
    f14_2.value = offer_2_home_inspection_check
    # f14_2.font = Font(bold=True)
    f14_2.alignment = Alignment(horizontal='center')
    f14_2.border = Border(top=hair, left=thin, right=thin)

    g14_2 = ws2['G14']
    g14_2.value = offer_3_home_inspection_check
    # g14_2.font = Font(bold=True)
    g14_2.alignment = Alignment(horizontal='center')
    g14_2.border = Border(top=hair, left=thin, right=thin)

    h14_2 = ws2['H14']
    h14_2.value = offer_4_home_inspection_check
    # h14_2.font = Font(bold=True)
    h14_2.alignment = Alignment(horizontal='center')
    h14_2.border = Border(top=hair, left=thin, right=thin)

    e15_2 = ws2['E15']
    e15_2.value = offer_2_home_inspection_days
    # e15_2.font = Font(bold=True)
    e15_2.alignment = Alignment(horizontal='center')
    e15_2.border = Border(bottom=hair, left=thin, right=thin)

    f15_2 = ws2['F15']
    f15_2.value = offer_2_home_inspection_days
    # f15_2.font = Font(bold=True)
    f15_2.alignment = Alignment(horizontal='center')
    f15_2.border = Border(bottom=hair, left=thin, right=thin)

    g15_2 = ws2['G15']
    g15_2.value = offer_3_home_inspection_days
    # g15_2.font = Font(bold=True)
    g15_2.alignment = Alignment(horizontal='center')
    g15_2.border = Border(bottom=hair, left=thin, right=thin)

    h15_2 = ws2['H15']
    h15_2.value = offer_4_home_inspection_days
    # h15_2.font = Font(bold=True)
    h15_2.alignment = Alignment(horizontal='center')
    h15_2.border = Border(bottom=hair, left=thin, right=thin)

    ws2.merge_cells('C16:C17')
    top_left_radon_insp_cont_2 = ws2['C16']
    top_left_radon_insp_cont_2.value = 'Radon Inspection Contingency'
    # top_left_radon_insp_cont_2.font = Font(bold=True)
    top_left_radon_insp_cont_2.alignment = Alignment(horizontal='left', vertical='center')
    top_left_radon_insp_cont_2.border = Border(top=hair, bottom=hair, left=thin)
    c16_2 = ws2['C16']
    c16_2.border = Border(top=hair, left=thin)
    d16_2 = ws2['D16']
    d16_2.border = Border(top=hair)
    c17_2 = ws2['C17']
    c17_2.border = Border(bottom=hair, left=thin)
    d17_2 = ws2['D17']
    d17_2.border = Border(bottom=hair)

    e16_2 = ws2['E16']
    e16_2.value = offer_2_radon_inspection_check
    # e16_2.font = Font(bold=True)
    e16_2.alignment = Alignment(horizontal='center')
    e16_2.border = Border(top=hair, left=thin, right=thin)

    f16_2 = ws2['F16']
    f16_2.value = offer_2_radon_inspection_check
    # f16_2.font = Font(bold=True)
    f16_2.alignment = Alignment(horizontal='center')
    f16_2.border = Border(top=hair, left=thin, right=thin)

    g16_2 = ws2['G16']
    g16_2.value = offer_3_radon_inspection_check
    # g16_2.font = Font(bold=True)
    g16_2.alignment = Alignment(horizontal='center')
    g16_2.border = Border(top=hair, left=thin, right=thin)

    h16_2 = ws2['H16']
    h16_2.value = offer_4_radon_inspection_check
    # h16_2.font = Font(bold=True)
    h16_2.alignment = Alignment(horizontal='center')
    h16_2.border = Border(top=hair, left=thin, right=thin)

    e17_2 = ws2['E17']
    e17_2.value = offer_2_radon_inspection_days
    # e17_2.font = Font(bold=True)
    e17_2.alignment = Alignment(horizontal='center')
    e17_2.border = Border(bottom=hair, left=thin, right=thin)

    f17_2 = ws2['F17']
    f17_2.value = offer_2_radon_inspection_days
    # f17_2.font = Font(bold=True)
    f17_2.alignment = Alignment(horizontal='center')
    f17_2.border = Border(bottom=hair, left=thin, right=thin)

    g17_2 = ws2['G17']
    g17_2.value = offer_3_radon_inspection_days
    # g17_2.font = Font(bold=True)
    g17_2.alignment = Alignment(horizontal='center')
    g17_2.border = Border(bottom=hair, left=thin, right=thin)

    h17_2 = ws2['H17']
    h17_2.value = offer_4_radon_inspection_days
    # h17_2.font = Font(bold=True)
    h17_2.alignment = Alignment(horizontal='center')
    h17_2.border = Border(bottom=hair, left=thin, right=thin)

    ws2.merge_cells('C18:C19')
    top_left_septic_insp_cont_2 = ws2['C18']
    top_left_septic_insp_cont_2.value = 'Septic Inspection Contingency'
    # top_left_septic_insp_cont_2.font = Font(bold=True)
    top_left_septic_insp_cont_2.alignment = Alignment(horizontal='left', vertical='center')
    top_left_septic_insp_cont_2.border = Border(top=hair, bottom=hair, left=thin)
    c18_2 = ws2['C18']
    c18_2.border = Border(top=hair, left=thin)
    d18_2 = ws2['D18']
    d18_2.border = Border(top=hair)
    c19_2 = ws2['C19']
    c19_2.border = Border(bottom=hair, left=thin)
    d19_2 = ws2['D19']
    d19_2.border = Border(bottom=hair)

    e18_2 = ws2['E18']
    e18_2.value = offer_2_septic_inspection_check
    # e18_2.font = Font(bold=True)
    e18_2.alignment = Alignment(horizontal='center')
    e18_2.border = Border(top=hair, left=thin, right=thin)

    f18_2 = ws2['F18']
    f18_2.value = offer_2_septic_inspection_check
    # f18_2.font = Font(bold=True)
    f18_2.alignment = Alignment(horizontal='center')
    f18_2.border = Border(top=hair, left=thin, right=thin)

    g18_2 = ws2['G18']
    g18_2.value = offer_3_septic_inspection_check
    # g18_2.font = Font(bold=True)
    g18_2.alignment = Alignment(horizontal='center')
    g18_2.border = Border(top=hair, left=thin, right=thin)

    h18_2 = ws2['H18']
    h18_2.value = offer_4_septic_inspection_check
    # h18_2.font = Font(bold=True)
    h18_2.alignment = Alignment(horizontal='center')
    h18_2.border = Border(top=hair, left=thin, right=thin)

    e19_2 = ws2['E19']
    e19_2.value = offer_2_septic_inspection_days
    # e19_2.font = Font(bold=True)
    e19_2.alignment = Alignment(horizontal='center')
    e19_2.border = Border(bottom=hair, left=thin, right=thin)

    f19_2 = ws2['F19']
    f19_2.value = offer_2_septic_inspection_days
    # f19_2.font = Font(bold=True)
    f19_2.alignment = Alignment(horizontal='center')
    f19_2.border = Border(bottom=hair, left=thin, right=thin)

    g19_2 = ws2['G19']
    g19_2.value = offer_3_septic_inspection_days
    # g19_2.font = Font(bold=True)
    g19_2.alignment = Alignment(horizontal='center')
    g19_2.border = Border(bottom=hair, left=thin, right=thin)

    h19_2 = ws2['H19']
    h19_2.value = offer_4_septic_inspection_days
    # h19_2.font = Font(bold=True)
    h19_2.alignment = Alignment(horizontal='center')
    h19_2.border = Border(bottom=hair, left=thin, right=thin)

    ws2.merge_cells('C20:C21')
    top_left_well_insp_cont_2 = ws2['C20']
    top_left_well_insp_cont_2.value = 'Well Inspection Contingency'
    # top_left_well_insp_cont_2.font = Font(bold=True)
    top_left_well_insp_cont_2.alignment = Alignment(horizontal='left', vertical='center')
    top_left_well_insp_cont_2.border = Border(top=hair, bottom=hair, left=thin)
    c20_2 = ws2['C20']
    c20_2.border = Border(top=hair, left=thin)
    d20_2 = ws2['D20']
    d20_2.border = Border(top=hair)
    c21_2 = ws2['C21']
    c21_2.border = Border(bottom=hair, left=thin)
    d21_2 = ws2['D21']
    d21_2.border = Border(bottom=hair)

    e20_2 = ws2['E20']
    e20_2.value = offer_2_well_inspection_check
    # e20_2.font = Font(bold=True)
    e20_2.alignment = Alignment(horizontal='center')
    e20_2.border = Border(top=hair, left=thin, right=thin)

    f20_2 = ws2['F20']
    f20_2.value = offer_2_well_inspection_check
    # f20_2.font = Font(bold=True)
    f20_2.alignment = Alignment(horizontal='center')
    f20_2.border = Border(top=hair, left=thin, right=thin)

    g20_2 = ws2['G20']
    g20_2.value = offer_3_well_inspection_check
    # g20_2.font = Font(bold=True)
    g20_2.alignment = Alignment(horizontal='center')
    g20_2.border = Border(top=hair, left=thin, right=thin)

    h20_2 = ws2['H20']
    h20_2.value = offer_4_well_inspection_check
    # h20_2.font = Font(bold=True)
    h20_2.alignment = Alignment(horizontal='center')
    h20_2.border = Border(top=hair, left=thin, right=thin)

    e21_2 = ws2['E21']
    e21_2.value = offer_2_well_inspection_days
    # e21_2.font = Font(bold=True)
    e21_2.alignment = Alignment(horizontal='center')
    e21_2.border = Border(bottom=hair, left=thin, right=thin)

    f21_2 = ws2['F21']
    f21_2.value = offer_2_well_inspection_days
    # f21_2.font = Font(bold=True)
    f21_2.alignment = Alignment(horizontal='center')
    f21_2.border = Border(bottom=hair, left=thin, right=thin)

    g21_2 = ws2['G21']
    g21_2.value = offer_3_well_inspection_days
    # g21_2.font = Font(bold=True)
    g21_2.alignment = Alignment(horizontal='center')
    g21_2.border = Border(bottom=hair, left=thin, right=thin)

    h21_2 = ws2['H21']
    h21_2.value = offer_4_well_inspection_days
    # h21_2.font = Font(bold=True)
    h21_2.alignment = Alignment(horizontal='center')
    h21_2.border = Border(bottom=hair, left=thin, right=thin)

    ws2.merge_cells('C22:C23')
    top_left_finance_cont_2 = ws2['C22']
    top_left_finance_cont_2.value = 'Finance Contingency'
    # top_left_finance_cont_2.font = Font(bold=True)
    top_left_finance_cont_2.alignment = Alignment(horizontal='left', vertical='center')
    top_left_finance_cont_2.border = Border(top=hair, bottom=hair, left=thin)
    c22_2 = ws2['C22']
    c22_2.border = Border(top=hair, left=thin)
    d22_2 = ws2['D22']
    d22_2.border = Border(top=hair)
    c23_2 = ws2['C23']
    c23_2.border = Border(bottom=hair, left=thin)
    d23_2 = ws2['D23']
    d23_2.border = Border(bottom=hair)

    e22_2 = ws2['E22']
    e22_2.value = offer_2_finance_contingency_check
    # e22_2.font = Font(bold=True)
    e22_2.alignment = Alignment(horizontal='center')
    e22_2.border = Border(top=hair, left=thin, right=thin)

    f22_2 = ws2['F22']
    f22_2.value = offer_2_finance_contingency_check
    # f22_2.font = Font(bold=True)
    f22_2.alignment = Alignment(horizontal='center')
    f22_2.border = Border(top=hair, left=thin, right=thin)

    g22_2 = ws2['G22']
    g22_2.value = offer_3_finance_contingency_check
    # g22_2.font = Font(bold=True)
    g22_2.alignment = Alignment(horizontal='center')
    g22_2.border = Border(top=hair, left=thin, right=thin)

    h22_2 = ws2['H22']
    h22_2.value = offer_4_finance_contingency_check
    # h22_2.font = Font(bold=True)
    h22_2.alignment = Alignment(horizontal='center')
    h22_2.border = Border(top=hair, left=thin, right=thin)

    e23_2 = ws2['E23']
    e23_2.value = offer_2_finance_contingency_days
    # e23_2.font = Font(bold=True)
    e23_2.alignment = Alignment(horizontal='center')
    e23_2.border = Border(bottom=hair, left=thin, right=thin)

    f23_2 = ws2['F23']
    f23_2.value = offer_2_finance_contingency_days
    # f23_2.font = Font(bold=True)
    f23_2.alignment = Alignment(horizontal='center')
    f23_2.border = Border(bottom=hair, left=thin, right=thin)

    g23_2 = ws2['G23']
    g23_2.value = offer_3_finance_contingency_days
    # g23_2.font = Font(bold=True)
    g23_2.alignment = Alignment(horizontal='center')
    g23_2.border = Border(bottom=hair, left=thin, right=thin)

    h23_2 = ws2['H23']
    h23_2.value = offer_4_finance_contingency_days
    # h23_2.font = Font(bold=True)
    h23_2.alignment = Alignment(horizontal='center')
    h23_2.border = Border(bottom=hair, left=thin, right=thin)

    ws2.merge_cells('C24:C25')
    top_left_appraisal_cont_2 = ws2['C24']
    top_left_appraisal_cont_2.value = 'Appraisal Contingency'
    # top_left_appraisal_cont_2.font = Font(bold=True)
    top_left_appraisal_cont_2.alignment = Alignment(horizontal='left', vertical='center')
    top_left_appraisal_cont_2.border = Border(top=hair, bottom=hair, left=thin)
    c24_2 = ws2['C24']
    c24_2.border = Border(top=hair, left=thin)
    d24_2 = ws2['D24']
    d24_2.border = Border(top=hair)
    c25_2 = ws2['C25']
    c25_2.border = Border(bottom=hair, left=thin)
    d25_2 = ws2['D25']
    d25_2.border = Border(bottom=hair)

    e24_2 = ws2['E24']
    e24_2.value = offer_2_appraisal_contingency_check
    # e24_2.font = Font(bold=True)
    e24_2.alignment = Alignment(horizontal='center')
    e24_2.border = Border(top=hair, left=thin, right=thin)

    f24_2 = ws2['F24']
    f24_2.value = offer_2_appraisal_contingency_check
    # f24_2.font = Font(bold=True)
    f24_2.alignment = Alignment(horizontal='center')
    f24_2.border = Border(top=hair, left=thin, right=thin)

    g24_2 = ws2['G24']
    g24_2.value = offer_3_appraisal_contingency_check
    # g24_2.font = Font(bold=True)
    g24_2.alignment = Alignment(horizontal='center')
    g24_2.border = Border(top=hair, left=thin, right=thin)

    h24_2 = ws2['H24']
    h24_2.value = offer_4_appraisal_contingency_check
    # h24_2.font = Font(bold=True)
    h24_2.alignment = Alignment(horizontal='center')
    h24_2.border = Border(top=hair, left=thin, right=thin)

    e25_2 = ws2['E25']
    e25_2.value = offer_2_appraisal_contingency_days
    # e25_2.font = Font(bold=True)
    e25_2.alignment = Alignment(horizontal='center')
    e25_2.border = Border(bottom=hair, left=thin, right=thin)

    f25_2 = ws2['F25']
    f25_2.value = offer_2_appraisal_contingency_days
    # f25_2.font = Font(bold=True)
    f25_2.alignment = Alignment(horizontal='center')
    f25_2.border = Border(bottom=hair, left=thin, right=thin)

    g25_2 = ws2['G25']
    g25_2.value = offer_3_appraisal_contingency_days
    # g25_2.font = Font(bold=True)
    g25_2.alignment = Alignment(horizontal='center')
    g25_2.border = Border(bottom=hair, left=thin, right=thin)

    h25_2 = ws2['H25']
    h25_2.value = offer_4_appraisal_contingency_days
    # h25_2.font = Font(bold=True)
    h25_2.alignment = Alignment(horizontal='center')
    h25_2.border = Border(bottom=hair, left=thin, right=thin)

    ws2.merge_cells('C26:C27')
    top_left_home_sale_cont_2 = ws2['C26']
    top_left_home_sale_cont_2.value = 'Home Sale Contingency'
    # top_left_home_sale_cont_2.font = Font(bold=True)
    top_left_home_sale_cont_2.alignment = Alignment(horizontal='left', vertical='center')
    top_left_home_sale_cont_2.border = Border(top=hair, bottom=hair, left=thin)
    c26_2 = ws2['C26']
    c26_2.border = Border(top=hair, left=thin)
    d26_2 = ws2['D26']
    d26_2.border = Border(top=hair)
    c27_2 = ws2['C27']
    c27_2.border = Border(bottom=hair, left=thin)
    d27_2 = ws2['D27']
    d27_2.border = Border(bottom=hair)

    e26_2 = ws2['E26']
    e26_2.value = offer_2_home_sale_contingency_check
    # e26_2.font = Font(bold=True)
    e26_2.alignment = Alignment(horizontal='center')
    e26_2.border = Border(top=hair, left=thin, right=thin)

    f26_2 = ws2['F26']
    f26_2.value = offer_2_home_sale_contingency_check
    # f26_2.font = Font(bold=True)
    f26_2.alignment = Alignment(horizontal='center')
    f26_2.border = Border(top=hair, left=thin, right=thin)

    g26_2 = ws2['G26']
    g26_2.value = offer_3_home_sale_contingency_check
    # g26_2.font = Font(bold=True)
    g26_2.alignment = Alignment(horizontal='center')
    g26_2.border = Border(top=hair, left=thin, right=thin)

    h26_2 = ws2['H26']
    h26_2.value = offer_4_home_sale_contingency_check
    # h26_2.font = Font(bold=True)
    h26_2.alignment = Alignment(horizontal='center')
    h26_2.border = Border(top=hair, left=thin, right=thin)

    e27_2 = ws2['E27']
    e27_2.value = offer_2_home_sale_contingency_days
    # e27_2.font = Font(bold=True)
    e27_2.alignment = Alignment(horizontal='center')
    e27_2.border = Border(bottom=hair, left=thin, right=thin)

    f27_2 = ws2['F27']
    f27_2.value = offer_2_home_sale_contingency_days
    # f27_2.font = Font(bold=True)
    f27_2.alignment = Alignment(horizontal='center')
    f27_2.border = Border(bottom=hair, left=thin, right=thin)

    g27_2 = ws2['G27']
    g27_2.value = offer_3_home_sale_contingency_days
    # g27_2.font = Font(bold=True)
    g27_2.alignment = Alignment(horizontal='center')
    g27_2.border = Border(bottom=hair, left=thin, right=thin)

    h27_2 = ws2['H27']
    h27_2.value = offer_4_home_sale_contingency_days
    # h27_2.font = Font(bold=True)
    h27_2.alignment = Alignment(horizontal='center')
    h27_2.border = Border(bottom=hair, left=thin, right=thin)

    c28_2 = ws2['C28']
    c28_2.value = 'Pre Occupancy Start Date'
    # c28_2.font = Font(bold=True)
    c28_2.alignment = Alignment(horizontal='left', vertical='center')
    c28_2.border = Border(top=hair, bottom=hair, left=thin)

    d28_2 = ws2['D28']
    d28_2.border = Border(top=hair, bottom=hair)

    e28_2 = ws2['E28']
    e28_2.value = offer_2_pre_occupancy_date
    # e28_2.font = Font(bold=True)
    e28_2.alignment = Alignment(horizontal='right')
    e28_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f28_2 = ws2['F28']
    f28_2.value = offer_2_pre_occupancy_date
    # f28_2.font = Font(bold=True)
    f28_2.alignment = Alignment(horizontal='right')
    f28_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g28_2 = ws2['G28']
    g28_2.value = offer_3_pre_occupancy_date
    # g28_2.font = Font(bold=True)
    g28_2.alignment = Alignment(horizontal='right')
    g28_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h28_2 = ws2['H28']
    h28_2.value = offer_4_pre_occupancy_date
    # h28_2.font = Font(bold=True)
    h28_2.alignment = Alignment(horizontal='right')
    h28_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c29_2 = ws2['C29']
    c29_2.value = 'Post Occupancy Thru Date'
    # c29_2.font = Font(bold=True)
    c29_2.alignment = Alignment(horizontal='left', vertical='center')
    c29_2.border = Border(top=hair, bottom=thin, left=thin)

    d29_2 = ws2['D29']
    d29_2.border = Border(top=hair, bottom=thin)

    e29_2 = ws2['E29']
    e29_2.value = offer_2_post_occupancy_date
    # e29_2.font = Font(bold=True)
    e29_2.alignment = Alignment(horizontal='right')
    e29_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f29_2 = ws2['F29']
    f29_2.value = offer_2_post_occupancy_date
    # f29_2.font = Font(bold=True)
    f29_2.alignment = Alignment(horizontal='right')
    f29_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g29_2 = ws2['G29']
    g29_2.value = offer_3_post_occupancy_date
    # g29_2.font = Font(bold=True)
    g29_2.alignment = Alignment(horizontal='right')
    g29_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h29_2 = ws2['H29']
    h29_2.value = offer_4_post_occupancy_date
    # h29_2.font = Font(bold=True)
    h29_2.alignment = Alignment(horizontal='right')
    h29_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    # Housing Related Cost Table
    c31_2 = ws2['C31']
    c31_2.value = 'HOUSING-RELATED COSTS'
    c31_2.font = Font(bold=True)
    c31_2.border = Border(bottom=thin)

    d31_2 = ws2['D31']
    d31_2.value = 'Calculation Description'
    d31_2.font = Font(bold=True)
    d31_2.border = Border(bottom=thin)

    c32_2 = ws2['C32']
    c32_2.value = 'Estimated Payoff - 1st Trust'
    c32_2.border = Border(top=thin, bottom=hair, left=thin)

    d32_2 = ws2['D32']
    d32_2.value = 'Principal Balance of Loan'
    d32_2.border = Border(top=thin, bottom=hair)

    e32_2 = ws2['E32']
    e32_2.value = first_trust
    e32_2.number_format = acct_fmt
    e32_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f32_2 = ws2['F32']
    f32_2.value = first_trust
    f32_2.number_format = acct_fmt
    f32_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g32_2 = ws2['G32']
    g32_2.value = first_trust
    g32_2.number_format = acct_fmt
    g32_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h32_2 = ws2['H32']
    h32_2.value = first_trust
    h32_2.number_format = acct_fmt
    h32_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c33_2 = ws2['C33']
    c33_2.value = 'Estimated Payoff - 2nd Trust'
    c33_2.border = Border(top=hair, bottom=hair, left=thin)

    d33_2 = ws2['D33']
    d33_2.value = 'Principal Balance of Loan'
    d33_2.border = Border(top=hair, bottom=hair)

    e33_2 = ws2['E33']
    e33_2.value = second_trust
    e33_2.number_format = acct_fmt
    e33_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f33_2 = ws2['F33']
    f33_2.value = second_trust
    f33_2.number_format = acct_fmt
    f33_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g33_2 = ws2['G33']
    g33_2.value = second_trust
    g33_2.number_format = acct_fmt
    g33_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h33_2 = ws2['H33']
    h33_2.value = second_trust
    h33_2.number_format = acct_fmt
    h33_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c34_2 = ws2['C34']
    c34_2.value = 'Purchaser Closing Cost / Contract'
    c34_2.border = Border(top=hair, bottom=hair, left=thin)

    d34_2 = ws2['D34']
    d34_2.value = 'Negotiated Into Contract'
    d34_2.border = Border(top=hair, bottom=hair)

    e34_2 = ws2['E34']
    e34_2.value = offer_2_closing_cost_subsidy_amt
    e34_2.number_format = acct_fmt
    e34_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f34_2 = ws2['F34']
    f34_2.value = offer_2_closing_cost_subsidy_amt
    f34_2.number_format = acct_fmt
    f34_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g34_2 = ws2['G34']
    g34_2.value = offer_3_closing_cost_subsidy_amt
    g34_2.number_format = acct_fmt
    g34_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h34_2 = ws2['H34']
    h34_2.value = offer_4_closing_cost_subsidy_amt
    h34_2.number_format = acct_fmt
    h34_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c35_2 = ws2['C35']
    c35_2.value = 'Prorated Taxes / Assessments'
    c35_2.border = Border(top=hair, bottom=hair, left=thin)

    d35_2 = ws2['D35']
    d35_2.value = '1 Year of Taxes Divided by 12 Multiplied by 3'
    d35_2.border = Border(top=hair, bottom=hair)

    e35_2 = ws2['E35']
    e35_2.value = prorated_taxes
    e35_2.number_format = acct_fmt
    e35_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f35_2 = ws2['F35']
    f35_2.value = prorated_taxes
    f35_2.number_format = acct_fmt
    f35_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g35_2 = ws2['G35']
    g35_2.value = prorated_taxes
    g35_2.number_format = acct_fmt
    g35_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h35_2 = ws2['H35']
    h35_2.value = prorated_taxes
    h35_2.number_format = acct_fmt
    h35_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c36_2 = ws2['C36']
    c36_2.value = 'Prorated HOA / Condo Dues'
    c36_2.border = Border(top=hair, bottom=thin, left=thin)

    d36_2 = ws2['D36']
    d36_2.value = '1 Year of Dues Divided by 12 Multiplied by 3'
    d36_2.border = Border(top=hair, bottom=thin)

    e36_2 = ws2['E36']
    e36_2.value = prorated_hoa_condo_fees
    e36_2.number_format = acct_fmt
    e36_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f36_2 = ws2['F36']
    f36_2.value = prorated_hoa_condo_fees
    f36_2.number_format = acct_fmt
    f36_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g36_2 = ws2['G36']
    g36_2.value = prorated_hoa_condo_fees
    g36_2.number_format = acct_fmt
    g36_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h36_2 = ws2['H36']
    h36_2.value = prorated_hoa_condo_fees
    h36_2.number_format = acct_fmt
    h36_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e37_2 = ws2['E37']
    e37_2.value = '=SUM(E32:E36)'
    e37_2.font = Font(bold=True)
    e37_2.number_format = acct_fmt
    e37_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f37_2 = ws2['F37']
    f37_2.value = '=SUM(F32:F36)'
    f37_2.font = Font(bold=True)
    f37_2.number_format = acct_fmt
    f37_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g37_2 = ws2['G37']
    g37_2.value = '=SUM(G32:G36)'
    g37_2.font = Font(bold=True)
    g37_2.number_format = acct_fmt
    g37_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h37_2 = ws2['H37']
    h37_2.value = '=SUM(H32:H36)'
    h37_2.font = Font(bold=True)
    h37_2.number_format = acct_fmt
    h37_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Broker and Financing Costs Table
    c38_2 = ws2['C38']
    c38_2.value = 'BROKERAGE & FINANCING COSTS'
    c38_2.font = Font(bold=True)
    c38_2.border = Border(bottom=thin)

    c39_2 = ws2['C39']
    c39_2.value = 'Listing Company Compensation'
    c39_2.border = Border(top=thin, bottom=hair, left=thin)

    d39_2 = ws2['D39']
    d39_2.value = '% from Listing Agreement * Offer Amount ($)'
    d39_2.border = Border(top=thin, bottom=hair)

    e39_2 = ws2['E39']
    e39_2.value = listing_company_pct * offer_2_amt
    e39_2.number_format = acct_fmt
    e39_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f39_2 = ws2['F39']
    f39_2.value = listing_company_pct * offer_2_amt
    f39_2.number_format = acct_fmt
    f39_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g39_2 = ws2['G39']
    g39_2.value = listing_company_pct * offer_3_amt
    g39_2.number_format = acct_fmt
    g39_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h39_2 = ws2['H39']
    h39_2.value = listing_company_pct * offer_4_amt
    h39_2.number_format = acct_fmt
    h39_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c40_2 = ws2['C40']
    c40_2.value = 'Selling Company Compensation'
    c40_2.border = Border(top=hair, bottom=hair, left=thin)

    d40_2 = ws2['D40']
    d40_2.value = '% from Listing Agreement * Offer Amount ($)'
    d40_2.border = Border(top=hair, bottom=hair)

    e40_2 = ws2['E40']
    e40_2.value = selling_company_pct * offer_2_amt
    e40_2.number_format = acct_fmt
    e40_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f40_2 = ws2['F40']
    f40_2.value = selling_company_pct * offer_2_amt
    f40_2.number_format = acct_fmt
    f40_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g40_2 = ws2['G40']
    g40_2.value = selling_company_pct * offer_3_amt
    g40_2.number_format = acct_fmt
    g40_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h40_2 = ws2['H40']
    h40_2.value = selling_company_pct * offer_4_amt
    h40_2.number_format = acct_fmt
    h40_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c41_2 = ws2['C41']
    c41_2.value = 'Processing Fee'
    c41_2.border = Border(top=hair, bottom=thin, left=thin)

    d41_2 = ws2['D41']
    d41_2.border = Border(top=hair, bottom=thin)

    e41_2 = ws2['E41']
    e41_2.value = processing_fee
    e41_2.number_format = acct_fmt
    e41_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f41_2 = ws2['F41']
    f41_2.value = processing_fee
    f41_2.number_format = acct_fmt
    f41_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g41_2 = ws2['G41']
    g41_2.value = processing_fee
    g41_2.number_format = acct_fmt
    g41_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h41_2 = ws2['H41']
    h41_2.value = processing_fee
    h41_2.number_format = acct_fmt
    h41_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e42_2 = ws2['E42']
    e42_2.value = '=SUM(E39:E41)'
    e42_2.font = Font(bold=True)
    e42_2.number_format = acct_fmt
    e42_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f42_2 = ws2['F42']
    f42_2.value = '=SUM(F39:F41)'
    f42_2.font = Font(bold=True)
    f42_2.number_format = acct_fmt
    f42_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g42_2 = ws2['G42']
    g42_2.value = '=SUM(G39:G41)'
    g42_2.font = Font(bold=True)
    g42_2.number_format = acct_fmt
    g42_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h42_2 = ws2['H42']
    h42_2.value = '=SUM(H39:H41)'
    h42_2.font = Font(bold=True)
    h42_2.number_format = acct_fmt
    h42_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Estimated Closing Costs Table
    c43_2 = ws2['C43']
    c43_2.value = 'ESTIMATED CLOSING COSTS'
    c43_2.font = Font(bold=True)
    c43_2.border = Border(bottom=thin)

    c44_2 = ws2['C44']
    c44_2.value = 'Settlement Fee'
    c44_2.border = Border(top=thin, bottom=hair, left=thin)

    d44_2 = ws2['D44']
    d44_2.value = 'Commonly Used Fee'
    d44_2.border = Border(top=thin, bottom=hair)

    e44_2 = ws2['E44']
    e44_2.value = settlement_fee
    e44_2.number_format = acct_fmt
    e44_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f44_2 = ws2['F44']
    f44_2.value = settlement_fee
    f44_2.number_format = acct_fmt
    f44_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g44_2 = ws2['G44']
    g44_2.value = settlement_fee
    g44_2.number_format = acct_fmt
    g44_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h44_2 = ws2['H44']
    h44_2.value = settlement_fee
    h44_2.number_format = acct_fmt
    h44_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c45_2 = ws2['C45']
    c45_2.value = 'Deed Preparation'
    c45_2.border = Border(top=hair, bottom=hair, left=thin)

    d45_2 = ws2['D45']
    d45_2.value = 'Commonly Used Fee'
    d45_2.border = Border(top=hair, bottom=hair)

    e45_2 = ws2['E45']
    e45_2.value = deed_preparation_fee
    e45_2.number_format = acct_fmt
    e45_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f45_2 = ws2['F45']
    f45_2.value = deed_preparation_fee
    f45_2.number_format = acct_fmt
    f45_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g45_2 = ws2['G45']
    g45_2.value = deed_preparation_fee
    g45_2.number_format = acct_fmt
    g45_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h45_2 = ws2['H45']
    h45_2.value = deed_preparation_fee
    h45_2.number_format = acct_fmt
    h45_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c46_2 = ws2['C46']
    c46_2.value = 'Release of Liens / Trusts'
    c46_2.border = Border(top=hair, bottom=thin, left=thin)

    d46_2 = ws2['D46']
    d46_2.value = 'Commonly Used Fee * Qty of Trusts or Liens'
    d46_2.border = Border(top=hair, bottom=thin)

    e46_2 = ws2['E46']
    e46_2.value = lien_trust_release_fee * lien_trust_release_qty
    e46_2.number_format = acct_fmt
    e46_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f46_2 = ws2['F46']
    f46_2.value = lien_trust_release_fee * lien_trust_release_qty
    f46_2.number_format = acct_fmt
    f46_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g46_2 = ws2['G46']
    g46_2.value = lien_trust_release_fee * lien_trust_release_qty
    g46_2.number_format = acct_fmt
    g46_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h46_2 = ws2['H46']
    h46_2.value = lien_trust_release_fee * lien_trust_release_qty
    h46_2.number_format = acct_fmt
    h46_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e47_2 = ws2['E47']
    e47_2.value = '=SUM(E44:E46)'
    e47_2.font = Font(bold=True)
    e47_2.number_format = acct_fmt
    e47_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f47_2 = ws2['F47']
    f47_2.value = '=SUM(F44:F46)'
    f47_2.font = Font(bold=True)
    f47_2.number_format = acct_fmt
    f47_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g47_2 = ws2['G47']
    g47_2.value = '=SUM(G44:G46)'
    g47_2.font = Font(bold=True)
    g47_2.number_format = acct_fmt
    g47_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h47_2 = ws2['H47']
    h47_2.value = '=SUM(H44:H46)'
    h47_2.font = Font(bold=True)
    h47_2.number_format = acct_fmt
    h47_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Misc Costs Table
    c48_2 = ws2['C48']
    c48_2.value = 'MISCELLANEOUS COSTS'
    c48_2.font = Font(bold=True)
    c48_2.border = Border(bottom=thin)

    c49_2 = ws2['C49']
    c49_2.value = 'Recording Release(s)'
    c49_2.border = Border(top=thin, bottom=hair, left=thin)

    d49_2 = ws2['D49']
    d49_2.value = 'Commonly Used Fee * Qty of Trusts Recorded'
    d49_2.border = Border(top=thin, bottom=hair)

    e49_2 = ws2['E49']
    e49_2.value = recording_fee * recording_trusts_liens_qty
    e49_2.number_format = acct_fmt
    e49_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f49_2 = ws2['F49']
    f49_2.value = recording_fee * recording_trusts_liens_qty
    f49_2.number_format = acct_fmt
    f49_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g49_2 = ws2['G49']
    g49_2.value = recording_fee * recording_trusts_liens_qty
    g49_2.number_format = acct_fmt
    g49_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h49_2 = ws2['H49']
    h49_2.value = recording_fee * recording_trusts_liens_qty
    h49_2.number_format = acct_fmt
    h49_2.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c50_2 = ws2['C50']
    c50_2.value = 'Grantor\'s Tax'
    c50_2.border = Border(top=hair, bottom=hair, left=thin)

    d50_2 = ws2['D50']
    d50_2.value = '% of Offer Amount ($)'
    d50_2.border = Border(top=hair, bottom=hair)

    e50_2 = ws2['E50']
    e50_2.value = grantors_tax_pct * offer_2_amt
    e50_2.number_format = acct_fmt
    e50_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f50_2 = ws2['F50']
    f50_2.value = grantors_tax_pct * offer_2_amt
    f50_2.number_format = acct_fmt
    f50_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g50_2 = ws2['G50']
    g50_2.value = grantors_tax_pct * offer_3_amt
    g50_2.number_format = acct_fmt
    g50_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h50_2 = ws2['H50']
    h50_2.value = grantors_tax_pct * offer_4_amt
    h50_2.number_format = acct_fmt
    h50_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c51_2 = ws2['C51']
    c51_2.value = 'Congestion Relief Tax'
    c51_2.border = Border(top=hair, bottom=hair, left=thin)

    d51_2 = ws2['D51']
    d51_2.value = '% of Offer Amount ($)'
    d51_2.border = Border(top=hair, bottom=hair)

    e51_2 = ws2['E51']
    e51_2.value = congestion_tax_pct * offer_2_amt
    e51_2.number_format = acct_fmt
    e51_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f51_2 = ws2['F51']
    f51_2.value = congestion_tax_pct * offer_2_amt
    f51_2.number_format = acct_fmt
    f51_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g51_2 = ws2['G51']
    g51_2.value = congestion_tax_pct * offer_3_amt
    g51_2.number_format = acct_fmt
    g51_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h51_2 = ws2['H51']
    h51_2.value = congestion_tax_pct * offer_4_amt
    h51_2.number_format = acct_fmt
    h51_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c52_2 = ws2['C52']
    c52_2.value = 'Pest Inspection'
    c52_2.border = Border(top=hair, bottom=hair, left=thin)

    d52_2 = ws2['D52']
    d52_2.value = 'Commonly Used Fee'
    d52_2.border = Border(top=hair, bottom=hair)

    e52_2 = ws2['E52']
    e52_2.value = pest_inspection_fee
    e52_2.number_format = acct_fmt
    e52_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f52_2 = ws2['F52']
    f52_2.value = pest_inspection_fee
    f52_2.number_format = acct_fmt
    f52_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g52_2 = ws2['G52']
    g52_2.value = pest_inspection_fee
    g52_2.number_format = acct_fmt
    g52_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h52_2 = ws2['H52']
    h52_2.value = pest_inspection_fee
    h52_2.number_format = acct_fmt
    h52_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c53_2 = ws2['C53']
    c53_2.value = 'POA / Condo Disclosures'
    c53_2.border = Border(top=hair, bottom=hair, left=thin)

    d53_2 = ws2['D53']
    d53_2.value = 'Commonly Used Fee'
    d53_2.border = Border(top=hair, bottom=hair)

    e53_2 = ws2['E53']
    e53_2.value = poa_condo_disclosure_fee
    e53_2.number_format = acct_fmt
    e53_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f53_2 = ws2['F53']
    f53_2.value = poa_condo_disclosure_fee
    f53_2.number_format = acct_fmt
    f53_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g53_2 = ws2['G53']
    g53_2.value = poa_condo_disclosure_fee
    g53_2.number_format = acct_fmt
    g53_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h53_2 = ws2['H53']
    h53_2.value = poa_condo_disclosure_fee
    h53_2.number_format = acct_fmt
    h53_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c54_2 = ws2['C54']
    c54_2.value = 'Pre Occupancy Credit to Seller'
    c54_2.border = Border(top=hair, bottom=hair, left=thin)

    d54_2 = ws2['D54']
    d54_2.value = 'Negotiated Into Contract'
    d54_2.border = Border(top=hair, bottom=hair)

    e54_2 = ws2['E54']
    e54_2.value = offer_2_pre_occupancy_credit_amt
    e54_2.number_format = acct_fmt
    e54_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f54_2 = ws2['F54']
    f54_2.value = offer_2_pre_occupancy_credit_amt
    f54_2.number_format = acct_fmt
    f54_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g54_2 = ws2['G54']
    g54_2.value = offer_3_pre_occupancy_credit_amt
    g54_2.number_format = acct_fmt
    g54_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h54_2 = ws2['H54']
    h54_2.value = offer_4_pre_occupancy_credit_amt
    h54_2.number_format = acct_fmt
    h54_2.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c55_2 = ws2['C55']
    c55_2.value = 'Post Occupancy Cost to Seller'
    c55_2.border = Border(top=hair, bottom=thin, left=thin)

    d55_2 = ws2['D55']
    d55_2.value = 'Negotiated Into Contract'
    d55_2.border = Border(top=hair, bottom=thin)

    e55_2 = ws2['E55']
    e55_2.value = offer_2_post_occupancy_cost_amt
    e55_2.number_format = acct_fmt
    e55_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f55_2 = ws2['F55']
    f55_2.value = offer_2_post_occupancy_cost_amt
    f55_2.number_format = acct_fmt
    f55_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g55_2 = ws2['G55']
    g55_2.value = offer_3_post_occupancy_cost_amt
    g55_2.number_format = acct_fmt
    g55_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h55_2 = ws2['H55']
    h55_2.value = offer_4_post_occupancy_cost_amt
    h55_2.number_format = acct_fmt
    h55_2.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e56_2 = ws2['E56']
    e56_2.value = '=SUM(E49:E53,E55)-E54'
    e56_2.font = Font(bold=True)
    e56_2.number_format = acct_fmt
    e56_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f56_2 = ws2['F56']
    f56_2.value = '=SUM(F49:F53,F55)-F54'
    f56_2.font = Font(bold=True)
    f56_2.number_format = acct_fmt
    f56_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g56_2 = ws2['G56']
    g56_2.value = '=SUM(G49:G53,G55)-G54'
    g56_2.font = Font(bold=True)
    g56_2.number_format = acct_fmt
    g56_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h56_2 = ws2['H56']
    h56_2.value = '=SUM(H49:H53,H55)-H54'
    h56_2.font = Font(bold=True)
    h56_2.number_format = acct_fmt
    h56_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Subtotals section
    ws2.merge_cells('C58:D58')
    top_left_cell_four_2 = ws2['C58']
    top_left_cell_four_2.value = 'TOTAL ESTIMATED COST OF SETTLEMENT'
    top_left_cell_four_2.font = Font(bold=True)
    top_left_cell_four_2.alignment = Alignment(horizontal='right')

    e58_2 = ws2['E58']
    e58_2.value = '=SUM(E37,E42,E47,E56)'
    e58_2.font = Font(bold=True)
    e58_2.number_format = acct_fmt
    e58_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f58_2 = ws2['F58']
    f58_2.value = '=SUM(F37,F42,F47,F56)'
    f58_2.font = Font(bold=True)
    f58_2.number_format = acct_fmt
    f58_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g58_2 = ws2['G58']
    g58_2.value = '=SUM(G37,G42,G47,G56)'
    g58_2.font = Font(bold=True)
    g58_2.number_format = acct_fmt
    g58_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h58_2 = ws2['H58']
    h58_2.value = '=SUM(H37,H42,H47,H56)'
    h58_2.font = Font(bold=True)
    h58_2.number_format = acct_fmt
    h58_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws2.merge_cells('C59:D59')
    top_left_cell_five_2 = ws2['C59']
    top_left_cell_five_2.value = 'Offer Amount ($)'
    top_left_cell_five_2.font = Font(bold=True)
    top_left_cell_five_2.alignment = Alignment(horizontal='right')

    e59_2 = ws2['E59']
    e59_2.value = offer_2_amt
    e59_2.font = Font(bold=True)
    e59_2.number_format = acct_fmt
    e59_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f59_2 = ws2['F59']
    f59_2.value = offer_2_amt
    f59_2.font = Font(bold=True)
    f59_2.number_format = acct_fmt
    f59_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g59_2 = ws2['G59']
    g59_2.value = offer_3_amt
    g59_2.font = Font(bold=True)
    g59_2.number_format = acct_fmt
    g59_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h59_2 = ws2['H59']
    h59_2.value = offer_4_amt
    h59_2.font = Font(bold=True)
    h59_2.number_format = acct_fmt
    h59_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws2.merge_cells('C60:D60')
    top_left_cell_six_2 = ws2['C60']
    top_left_cell_six_2.value = 'LESS: Total Estimated Cost of Settlement'
    top_left_cell_six_2.font = Font(bold=True)
    top_left_cell_six_2.alignment = Alignment(horizontal='right')

    e60_2 = ws2['E60']
    e60_2.value = '=-SUM(E37,E42,E47,E56)'
    e60_2.font = Font(bold=True)
    e60_2.number_format = acct_fmt
    e60_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f60_2 = ws2['F60']
    f60_2.value = '=-SUM(F37,F42,F47,F56)'
    f60_2.font = Font(bold=True)
    f60_2.number_format = acct_fmt
    f60_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g60_2 = ws2['G60']
    g60_2.value = '=-SUM(G37,G42,G47,G56)'
    g60_2.font = Font(bold=True)
    g60_2.number_format = acct_fmt
    g60_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h60_2 = ws2['H60']
    h60_2.value = '=-SUM(H37,H42,H47,H56)'
    h60_2.font = Font(bold=True)
    h60_2.number_format = acct_fmt
    h60_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws2.merge_cells('C61:D61')
    top_left_cell_seven_2 = ws2['C61']
    top_left_cell_seven_2.value = 'ESTIMATED TOTAL NET PROCEEDS'
    top_left_cell_seven_2.font = Font(bold=True)
    top_left_cell_seven_2.alignment = Alignment(horizontal='right')

    e61_2 = ws2['E61']
    e61_2.value = '=SUM(E59:E60)'
    e61_2.font = Font(bold=True)
    e61_2.number_format = acct_fmt
    e61_2.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    e61_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f61_2 = ws2['F61']
    f61_2.value = '=SUM(F59:F60)'
    f61_2.font = Font(bold=True)
    f61_2.number_format = acct_fmt
    f61_2.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    f61_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g61_2 = ws2['G61']
    g61_2.value = '=SUM(G59:G60)'
    g61_2.font = Font(bold=True)
    g61_2.number_format = acct_fmt
    g61_2.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    g61_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h61_2 = ws2['H61']
    h61_2.value = '=SUM(H59:H60)'
    h61_2.font = Font(bold=True)
    h61_2.number_format = acct_fmt
    h61_2.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    h61_2.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Signature Block
    c63_2 = ws2['C63']
    c63_2.value = 'PREPARED BY:'

    c64_2 = ws2['C64']
    c64_2.value = agent

    e63_2 = ws2['E63']
    e63_2.value = 'SELLER:'

    e64_2 = ws2['E64']
    e64_2.value = seller_name

    # Freedom Logo
    # c53 = ws2['C53']
    freedom_logo = drawing.image.Image('freedom_logo.png')
    freedom_logo.height = 80
    freedom_logo.width = 115
    freedom_logo.anchor = 'C66'
    ws2.add_image(freedom_logo)

    # Disclosure Statement
    ws2.merge_cells('C70:H74')
    top_left_cell_eight_2 = ws2['C70']
    top_left_cell_eight_2.value = '''These estimates are not guaranteed and may not include escrows. Escrow balances are reimbursed by the existing lender. Taxes, rents & association dues are pro-rated at settlement. Under Virginia Law, the seller's proceeds may not be available for up to 2 business days following recording of the deed. Seller acknowledges receipt of this statement.'''
    top_left_cell_eight_2.font = Font(italic=True)
    top_left_cell_eight_2.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

    # page 3
    # Build Header
    ws3.merge_cells('C2:H2')
    top_left_cell_one_3 = ws3['C2']
    top_left_cell_one_3.value = 'Seller\'s Total Net Proceeds For Different Offers'
    top_left_cell_one_3.font = Font(bold=True)
    top_left_cell_one_3.alignment = Alignment(horizontal='center')

    ws3.merge_cells('C3:H3')
    top_left_cell_two_3 = ws3['C3']
    top_left_cell_two_3.value = f'{seller_name} - {seller_address}'
    top_left_cell_two_3.font = Font(bold=True)
    top_left_cell_two_3.alignment = Alignment(horizontal='center')

    ws3.merge_cells('C4:H4')
    top_left_cell_three_3 = ws3['C4']
    top_left_cell_three_3.value = f'Date Prepared: {date}'
    top_left_cell_three_3.font = Font(bold=True)
    top_left_cell_three_3.alignment = Alignment(horizontal='center')

    ws3.merge_cells('C5:H5')
    top_left_cell_four_3 = ws3['C5']
    top_left_cell_four_3.value = f'List Price: ${list_price:,.2f}'
    top_left_cell_four_3.font = Font(bold=True)
    top_left_cell_four_3.alignment = Alignment(horizontal='center')

    c7_3 = ws3['C7']
    c7_3.value = 'OFFER SUMMARY FEATURES'
    c7_3.font = Font(bold=True)
    c7_3.border = Border(bottom=thin)

    e7_3 = ws3['E7']
    e7_3.value = offer_3_name
    e7_3.font = Font(bold=True)
    e7_3.border = Border(bottom=thin)
    e7_3.alignment = Alignment(horizontal='center', wrap_text=True)

    f7_3 = ws3['F7']
    f7_3.value = offer_3_name
    f7_3.font = Font(bold=True)
    f7_3.border = Border(bottom=thin)
    f7_3.alignment = Alignment(horizontal='center', wrap_text=True)

    g7_3 = ws3['G7']
    g7_3.value = offer_3_name
    g7_3.font = Font(bold=True)
    g7_3.border = Border(bottom=thin)
    g7_3.alignment = Alignment(horizontal='center', wrap_text=True)

    h7_3 = ws3['H7']
    h7_3.value = offer_4_name
    h7_3.font = Font(bold=True)
    h7_3.border = Border(bottom=thin)
    h7_3.alignment = Alignment(horizontal='center', wrap_text=True)

    c8_3 = ws3['C8']
    c8_3.value = 'Offer Amt. ($)'
    # c8_3.font = Font(bold=True)
    c8_3.alignment = Alignment(horizontal='left', vertical='center')
    c8_3.border = Border(top=thin, bottom=hair, left=thin)

    d8_3 = ws3['D8']
    d8_3.border = Border(top=thin, bottom=hair)

    e8_3 = ws3['E8']
    e8_3.value = offer_3_amt
    # e8_3.font = Font(bold=True)
    e8_3.number_format = acct_fmt
    e8_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f8_3 = ws3['F8']
    f8_3.value = offer_3_amt
    # f8_3.font = Font(bold=True)
    f8_3.number_format = acct_fmt
    f8_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g8_3 = ws3['G8']
    g8_3.value = offer_3_amt
    # g8_3.font = Font(bold=True)
    g8_3.number_format = acct_fmt
    g8_3.border = Border(top=thin, bottom=hair, right=thin, left=thin)

    h8_3 = ws3['H8']
    h8_3.value = offer_4_amt
    # h8_3.font = Font(bold=True)
    h8_3.number_format = acct_fmt
    h8_3.border = Border(top=thin, bottom=hair, right=thin, left=thin)

    c9_3 = ws3['C9']
    c9_3.value = 'Down Pmt (%)'
    # c9_3.font = Font(bold=True)
    c9_3.alignment = Alignment(horizontal='left', vertical='center')
    c9_3.border = Border(top=hair, bottom=hair, left=thin)

    d9_3 = ws3['D9']
    d9_3.border = Border(top=hair, bottom=hair)

    e9_3 = ws3['E9']
    e9_3.value = offer_3_down_pmt_pct
    # e9_3.font = Font(bold=True)
    e9_3.number_format = pct_fmt
    e9_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f9_3 = ws3['F9']
    f9_3.value = offer_3_down_pmt_pct
    # f9_3.font = Font(bold=True)
    f9_3.number_format = pct_fmt
    f9_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g9_3 = ws3['G9']
    g9_3.value = offer_3_down_pmt_pct
    # g9_3.font = Font(bold=True)
    g9_3.number_format = pct_fmt
    g9_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h9_3 = ws3['H9']
    h9_3.value = offer_4_down_pmt_pct
    # h9_3.font = Font(bold=True)
    h9_3.number_format = pct_fmt
    h9_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c10_3 = ws3['C10']
    c10_3.value = 'Settlement Date'
    # c10_3.font = Font(bold=True)
    c10_3.alignment = Alignment(horizontal='left', vertical='center')
    c10_3.border = Border(top=hair, bottom=hair, left=thin)

    d10_3 = ws3['D10']
    d10_3.border = Border(top=hair, bottom=hair)

    e10_3 = ws3['E10']
    e10_3.value = offer_3_settlement_date
    # e10_3.font = Font(bold=True)
    e10_3.alignment = Alignment(horizontal='right')
    e10_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f10_3 = ws3['F10']
    f10_3.value = offer_3_settlement_date
    # f10_3.font = Font(bold=True)
    f10_3.alignment = Alignment(horizontal='right')
    f10_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g10_3 = ws3['G10']
    g10_3.value = offer_3_settlement_date
    # g10_3.font = Font(bold=True)
    g10_3.alignment = Alignment(horizontal='right')
    g10_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h10_3 = ws3['H10']
    h10_3.value = offer_4_settlement_date
    # h10_3.font = Font(bold=True)
    h10_3.alignment = Alignment(horizontal='right')
    h10_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c11_3 = ws3['C11']
    c11_3.value = 'Settlement Company'
    # c11_3.font = Font(bold=True)
    c11_3.alignment = Alignment(horizontal='left', vertical='center')
    c11_3.border = Border(top=hair, bottom=hair, left=thin)

    d11_3 = ws3['D11']
    d11_3.border = Border(top=hair, bottom=hair)

    e11_3 = ws3['E11']
    e11_3.value = offer_3_settlement_company
    # e11_3.font = Font(bold=True)
    e11_3.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    e11_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f11_3 = ws3['F11']
    f11_3.value = offer_3_settlement_company
    # f11_3.font = Font(bold=True)
    f11_3.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    f11_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g11_3 = ws3['G11']
    g11_3.value = offer_3_settlement_company
    # g11_3.font = Font(bold=True)
    g11_3.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    g11_3.border = Border(top=hair, bottom=hair, right=thin, left=thin)

    h11_3 = ws3['H11']
    h11_3.value = offer_4_settlement_company
    # h11_3.font = Font(bold=True)
    h11_3.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    h11_3.border = Border(top=hair, bottom=hair, right=thin, left=thin)

    c12_3 = ws3['C12']
    c12_3.value = 'EMD Amt. ($)'
    # c12_3.font = Font(bold=True)
    c12_3.alignment = Alignment(horizontal='left', vertical='center')
    c12_3.border = Border(top=hair, bottom=hair, left=thin)

    d12_3 = ws3['D12']
    d12_3.border = Border(top=hair, bottom=hair)

    e12_3 = ws3['E12']
    e12_3.value = offer_3_emd_amt
    # e12_3.font = Font(bold=True)
    e12_3.number_format = acct_fmt
    e12_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f12_3 = ws3['F12']
    f12_3.value = offer_3_emd_amt
    # f12_3.font = Font(bold=True)
    f12_3.number_format = acct_fmt
    f12_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g12_3 = ws3['G12']
    g12_3.value = offer_3_emd_amt
    # g12_3.font = Font(bold=True)
    g12_3.number_format = acct_fmt
    g12_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h12_3 = ws3['H12']
    h12_3.value = offer_4_emd_amt
    # h12_3.font = Font(bold=True)
    h12_3.number_format = acct_fmt
    h12_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c13_3 = ws3['C13']
    c13_3.value = 'Financing Type'
    # c13_3.font = Font(bold=True)
    c13_3.alignment = Alignment(horizontal='left', vertical='center')
    c13_3.border = Border(top=hair, bottom=hair, left=thin)

    d13_3 = ws3['D13']
    d13_3.border = Border(top=hair, bottom=hair)

    e13_3 = ws3['E13']
    e13_3.value = offer_3_financing_type
    # e13_3.font = Font(bold=True)
    e13_3.alignment = Alignment(horizontal='center')
    e13_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f13_3 = ws3['F13']
    f13_3.value = offer_3_financing_type
    # f13_3.font = Font(bold=True)
    f13_3.alignment = Alignment(horizontal='center')
    f13_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g13_3 = ws3['G13']
    g13_3.value = offer_3_financing_type
    # g13_3.font = Font(bold=True)
    g13_3.alignment = Alignment(horizontal='center')
    g13_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h13_3 = ws3['H13']
    h13_3.value = offer_4_financing_type
    # h13_3.font = Font(bold=True)
    h13_3.alignment = Alignment(horizontal='center')
    h13_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    ws3.merge_cells('C14:C15')
    top_left_home_insp_cont_3 = ws3['C14']
    top_left_home_insp_cont_3.value = 'Home Inspection Contingency'
    # top_left_home_insp_cont_3.font = Font(bold=True)
    top_left_home_insp_cont_3.alignment = Alignment(horizontal='left', vertical='center')
    top_left_home_insp_cont_3.border = Border(top=hair, bottom=hair, left=thin)
    c14_3 = ws3['C14']
    c14_3.border = Border(top=hair, left=thin)
    d14_3 = ws3['D14']
    d14_3.border = Border(top=hair)
    c15_3 = ws3['C15']
    c15_3.border = Border(bottom=hair, left=thin)
    d15_3 = ws3['D15']
    d15_3.border = Border(bottom=hair)

    e14_3 = ws3['E14']
    e14_3.value = offer_3_home_inspection_check
    # e14_3.font = Font(bold=True)
    e14_3.alignment = Alignment(horizontal='center')
    e14_3.border = Border(top=hair, left=thin, right=thin)

    f14_3 = ws3['F14']
    f14_3.value = offer_3_home_inspection_check
    # f14_3.font = Font(bold=True)
    f14_3.alignment = Alignment(horizontal='center')
    f14_3.border = Border(top=hair, left=thin, right=thin)

    g14_3 = ws3['G14']
    g14_3.value = offer_3_home_inspection_check
    # g14_3.font = Font(bold=True)
    g14_3.alignment = Alignment(horizontal='center')
    g14_3.border = Border(top=hair, left=thin, right=thin)

    h14_3 = ws3['H14']
    h14_3.value = offer_4_home_inspection_check
    # h14_3.font = Font(bold=True)
    h14_3.alignment = Alignment(horizontal='center')
    h14_3.border = Border(top=hair, left=thin, right=thin)

    e15_3 = ws3['E15']
    e15_3.value = offer_3_home_inspection_days
    # e15_3.font = Font(bold=True)
    e15_3.alignment = Alignment(horizontal='center')
    e15_3.border = Border(bottom=hair, left=thin, right=thin)

    f15_3 = ws3['F15']
    f15_3.value = offer_3_home_inspection_days
    # f15_3.font = Font(bold=True)
    f15_3.alignment = Alignment(horizontal='center')
    f15_3.border = Border(bottom=hair, left=thin, right=thin)

    g15_3 = ws3['G15']
    g15_3.value = offer_3_home_inspection_days
    # g15_3.font = Font(bold=True)
    g15_3.alignment = Alignment(horizontal='center')
    g15_3.border = Border(bottom=hair, left=thin, right=thin)

    h15_3 = ws3['H15']
    h15_3.value = offer_4_home_inspection_days
    # h15_3.font = Font(bold=True)
    h15_3.alignment = Alignment(horizontal='center')
    h15_3.border = Border(bottom=hair, left=thin, right=thin)

    ws3.merge_cells('C16:C17')
    top_left_radon_insp_cont_3 = ws3['C16']
    top_left_radon_insp_cont_3.value = 'Radon Inspection Contingency'
    # top_left_radon_insp_cont_3.font = Font(bold=True)
    top_left_radon_insp_cont_3.alignment = Alignment(horizontal='left', vertical='center')
    top_left_radon_insp_cont_3.border = Border(top=hair, bottom=hair, left=thin)
    c16_3 = ws3['C16']
    c16_3.border = Border(top=hair, left=thin)
    d16_3 = ws3['D16']
    d16_3.border = Border(top=hair)
    c17_3 = ws3['C17']
    c17_3.border = Border(bottom=hair, left=thin)
    d17_3 = ws3['D17']
    d17_3.border = Border(bottom=hair)

    e16_3 = ws3['E16']
    e16_3.value = offer_3_radon_inspection_check
    # e16_3.font = Font(bold=True)
    e16_3.alignment = Alignment(horizontal='center')
    e16_3.border = Border(top=hair, left=thin, right=thin)

    f16_3 = ws3['F16']
    f16_3.value = offer_3_radon_inspection_check
    # f16_3.font = Font(bold=True)
    f16_3.alignment = Alignment(horizontal='center')
    f16_3.border = Border(top=hair, left=thin, right=thin)

    g16_3 = ws3['G16']
    g16_3.value = offer_3_radon_inspection_check
    # g16_3.font = Font(bold=True)
    g16_3.alignment = Alignment(horizontal='center')
    g16_3.border = Border(top=hair, left=thin, right=thin)

    h16_3 = ws3['H16']
    h16_3.value = offer_4_radon_inspection_check
    # h16_3.font = Font(bold=True)
    h16_3.alignment = Alignment(horizontal='center')
    h16_3.border = Border(top=hair, left=thin, right=thin)

    e17_3 = ws3['E17']
    e17_3.value = offer_3_radon_inspection_days
    # e17_3.font = Font(bold=True)
    e17_3.alignment = Alignment(horizontal='center')
    e17_3.border = Border(bottom=hair, left=thin, right=thin)

    f17_3 = ws3['F17']
    f17_3.value = offer_3_radon_inspection_days
    # f17_3.font = Font(bold=True)
    f17_3.alignment = Alignment(horizontal='center')
    f17_3.border = Border(bottom=hair, left=thin, right=thin)

    g17_3 = ws3['G17']
    g17_3.value = offer_3_radon_inspection_days
    # g17_3.font = Font(bold=True)
    g17_3.alignment = Alignment(horizontal='center')
    g17_3.border = Border(bottom=hair, left=thin, right=thin)

    h17_3 = ws3['H17']
    h17_3.value = offer_4_radon_inspection_days
    # h17_3.font = Font(bold=True)
    h17_3.alignment = Alignment(horizontal='center')
    h17_3.border = Border(bottom=hair, left=thin, right=thin)

    ws3.merge_cells('C18:C19')
    top_left_septic_insp_cont_3 = ws3['C18']
    top_left_septic_insp_cont_3.value = 'Septic Inspection Contingency'
    # top_left_septic_insp_cont_3.font = Font(bold=True)
    top_left_septic_insp_cont_3.alignment = Alignment(horizontal='left', vertical='center')
    top_left_septic_insp_cont_3.border = Border(top=hair, bottom=hair, left=thin)
    c18_3 = ws3['C18']
    c18_3.border = Border(top=hair, left=thin)
    d18_3 = ws3['D18']
    d18_3.border = Border(top=hair)
    c19_3 = ws3['C19']
    c19_3.border = Border(bottom=hair, left=thin)
    d19_3 = ws3['D19']
    d19_3.border = Border(bottom=hair)

    e18_3 = ws3['E18']
    e18_3.value = offer_3_septic_inspection_check
    # e18_3.font = Font(bold=True)
    e18_3.alignment = Alignment(horizontal='center')
    e18_3.border = Border(top=hair, left=thin, right=thin)

    f18_3 = ws3['F18']
    f18_3.value = offer_3_septic_inspection_check
    # f18_3.font = Font(bold=True)
    f18_3.alignment = Alignment(horizontal='center')
    f18_3.border = Border(top=hair, left=thin, right=thin)

    g18_3 = ws3['G18']
    g18_3.value = offer_3_septic_inspection_check
    # g18_3.font = Font(bold=True)
    g18_3.alignment = Alignment(horizontal='center')
    g18_3.border = Border(top=hair, left=thin, right=thin)

    h18_3 = ws3['H18']
    h18_3.value = offer_4_septic_inspection_check
    # h18_3.font = Font(bold=True)
    h18_3.alignment = Alignment(horizontal='center')
    h18_3.border = Border(top=hair, left=thin, right=thin)

    e19_3 = ws3['E19']
    e19_3.value = offer_3_septic_inspection_days
    # e19_3.font = Font(bold=True)
    e19_3.alignment = Alignment(horizontal='center')
    e19_3.border = Border(bottom=hair, left=thin, right=thin)

    f19_3 = ws3['F19']
    f19_3.value = offer_3_septic_inspection_days
    # f19_3.font = Font(bold=True)
    f19_3.alignment = Alignment(horizontal='center')
    f19_3.border = Border(bottom=hair, left=thin, right=thin)

    g19_3 = ws3['G19']
    g19_3.value = offer_3_septic_inspection_days
    # g19_3.font = Font(bold=True)
    g19_3.alignment = Alignment(horizontal='center')
    g19_3.border = Border(bottom=hair, left=thin, right=thin)

    h19_3 = ws3['H19']
    h19_3.value = offer_4_septic_inspection_days
    # h19_3.font = Font(bold=True)
    h19_3.alignment = Alignment(horizontal='center')
    h19_3.border = Border(bottom=hair, left=thin, right=thin)

    ws3.merge_cells('C20:C21')
    top_left_well_insp_cont_3 = ws3['C20']
    top_left_well_insp_cont_3.value = 'Well Inspection Contingency'
    # top_left_well_insp_cont_3.font = Font(bold=True)
    top_left_well_insp_cont_3.alignment = Alignment(horizontal='left', vertical='center')
    top_left_well_insp_cont_3.border = Border(top=hair, bottom=hair, left=thin)
    c20_3 = ws3['C20']
    c20_3.border = Border(top=hair, left=thin)
    d20_3 = ws3['D20']
    d20_3.border = Border(top=hair)
    c21_3 = ws3['C21']
    c21_3.border = Border(bottom=hair, left=thin)
    d21_3 = ws3['D21']
    d21_3.border = Border(bottom=hair)

    e20_3 = ws3['E20']
    e20_3.value = offer_3_well_inspection_check
    # e20_3.font = Font(bold=True)
    e20_3.alignment = Alignment(horizontal='center')
    e20_3.border = Border(top=hair, left=thin, right=thin)

    f20_3 = ws3['F20']
    f20_3.value = offer_3_well_inspection_check
    # f20_3.font = Font(bold=True)
    f20_3.alignment = Alignment(horizontal='center')
    f20_3.border = Border(top=hair, left=thin, right=thin)

    g20_3 = ws3['G20']
    g20_3.value = offer_3_well_inspection_check
    # g20_3.font = Font(bold=True)
    g20_3.alignment = Alignment(horizontal='center')
    g20_3.border = Border(top=hair, left=thin, right=thin)

    h20_3 = ws3['H20']
    h20_3.value = offer_4_well_inspection_check
    # h20_3.font = Font(bold=True)
    h20_3.alignment = Alignment(horizontal='center')
    h20_3.border = Border(top=hair, left=thin, right=thin)

    e21_3 = ws3['E21']
    e21_3.value = offer_3_well_inspection_days
    # e21_3.font = Font(bold=True)
    e21_3.alignment = Alignment(horizontal='center')
    e21_3.border = Border(bottom=hair, left=thin, right=thin)

    f21_3 = ws3['F21']
    f21_3.value = offer_3_well_inspection_days
    # f21_3.font = Font(bold=True)
    f21_3.alignment = Alignment(horizontal='center')
    f21_3.border = Border(bottom=hair, left=thin, right=thin)

    g21_3 = ws3['G21']
    g21_3.value = offer_3_well_inspection_days
    # g21_3.font = Font(bold=True)
    g21_3.alignment = Alignment(horizontal='center')
    g21_3.border = Border(bottom=hair, left=thin, right=thin)

    h21_3 = ws3['H21']
    h21_3.value = offer_4_well_inspection_days
    # h21_3.font = Font(bold=True)
    h21_3.alignment = Alignment(horizontal='center')
    h21_3.border = Border(bottom=hair, left=thin, right=thin)

    ws3.merge_cells('C22:C23')
    top_left_finance_cont_3 = ws3['C22']
    top_left_finance_cont_3.value = 'Finance Contingency'
    # top_left_finance_cont_3.font = Font(bold=True)
    top_left_finance_cont_3.alignment = Alignment(horizontal='left', vertical='center')
    top_left_finance_cont_3.border = Border(top=hair, bottom=hair, left=thin)
    c22_3 = ws3['C22']
    c22_3.border = Border(top=hair, left=thin)
    d22_3 = ws3['D22']
    d22_3.border = Border(top=hair)
    c23_3 = ws3['C23']
    c23_3.border = Border(bottom=hair, left=thin)
    d23_3 = ws3['D23']
    d23_3.border = Border(bottom=hair)

    e22_3 = ws3['E22']
    e22_3.value = offer_3_finance_contingency_check
    # e22_3.font = Font(bold=True)
    e22_3.alignment = Alignment(horizontal='center')
    e22_3.border = Border(top=hair, left=thin, right=thin)

    f22_3 = ws3['F22']
    f22_3.value = offer_3_finance_contingency_check
    # f22_3.font = Font(bold=True)
    f22_3.alignment = Alignment(horizontal='center')
    f22_3.border = Border(top=hair, left=thin, right=thin)

    g22_3 = ws3['G22']
    g22_3.value = offer_3_finance_contingency_check
    # g22_3.font = Font(bold=True)
    g22_3.alignment = Alignment(horizontal='center')
    g22_3.border = Border(top=hair, left=thin, right=thin)

    h22_3 = ws3['H22']
    h22_3.value = offer_4_finance_contingency_check
    # h22_3.font = Font(bold=True)
    h22_3.alignment = Alignment(horizontal='center')
    h22_3.border = Border(top=hair, left=thin, right=thin)

    e23_3 = ws3['E23']
    e23_3.value = offer_3_finance_contingency_days
    # e23_3.font = Font(bold=True)
    e23_3.alignment = Alignment(horizontal='center')
    e23_3.border = Border(bottom=hair, left=thin, right=thin)

    f23_3 = ws3['F23']
    f23_3.value = offer_3_finance_contingency_days
    # f23_3.font = Font(bold=True)
    f23_3.alignment = Alignment(horizontal='center')
    f23_3.border = Border(bottom=hair, left=thin, right=thin)

    g23_3 = ws3['G23']
    g23_3.value = offer_3_finance_contingency_days
    # g23_3.font = Font(bold=True)
    g23_3.alignment = Alignment(horizontal='center')
    g23_3.border = Border(bottom=hair, left=thin, right=thin)

    h23_3 = ws3['H23']
    h23_3.value = offer_4_finance_contingency_days
    # h23_3.font = Font(bold=True)
    h23_3.alignment = Alignment(horizontal='center')
    h23_3.border = Border(bottom=hair, left=thin, right=thin)

    ws3.merge_cells('C24:C25')
    top_left_appraisal_cont_3 = ws3['C24']
    top_left_appraisal_cont_3.value = 'Appraisal Contingency'
    # top_left_appraisal_cont_3.font = Font(bold=True)
    top_left_appraisal_cont_3.alignment = Alignment(horizontal='left', vertical='center')
    top_left_appraisal_cont_3.border = Border(top=hair, bottom=hair, left=thin)
    c24_3 = ws3['C24']
    c24_3.border = Border(top=hair, left=thin)
    d24_3 = ws3['D24']
    d24_3.border = Border(top=hair)
    c25_3 = ws3['C25']
    c25_3.border = Border(bottom=hair, left=thin)
    d25_3 = ws3['D25']
    d25_3.border = Border(bottom=hair)

    e24_3 = ws3['E24']
    e24_3.value = offer_3_appraisal_contingency_check
    # e24_3.font = Font(bold=True)
    e24_3.alignment = Alignment(horizontal='center')
    e24_3.border = Border(top=hair, left=thin, right=thin)

    f24_3 = ws3['F24']
    f24_3.value = offer_3_appraisal_contingency_check
    # f24_3.font = Font(bold=True)
    f24_3.alignment = Alignment(horizontal='center')
    f24_3.border = Border(top=hair, left=thin, right=thin)

    g24_3 = ws3['G24']
    g24_3.value = offer_3_appraisal_contingency_check
    # g24_3.font = Font(bold=True)
    g24_3.alignment = Alignment(horizontal='center')
    g24_3.border = Border(top=hair, left=thin, right=thin)

    h24_3 = ws3['H24']
    h24_3.value = offer_4_appraisal_contingency_check
    # h24_3.font = Font(bold=True)
    h24_3.alignment = Alignment(horizontal='center')
    h24_3.border = Border(top=hair, left=thin, right=thin)

    e25_3 = ws3['E25']
    e25_3.value = offer_3_appraisal_contingency_days
    # e25_3.font = Font(bold=True)
    e25_3.alignment = Alignment(horizontal='center')
    e25_3.border = Border(bottom=hair, left=thin, right=thin)

    f25_3 = ws3['F25']
    f25_3.value = offer_3_appraisal_contingency_days
    # f25_3.font = Font(bold=True)
    f25_3.alignment = Alignment(horizontal='center')
    f25_3.border = Border(bottom=hair, left=thin, right=thin)

    g25_3 = ws3['G25']
    g25_3.value = offer_3_appraisal_contingency_days
    # g25_3.font = Font(bold=True)
    g25_3.alignment = Alignment(horizontal='center')
    g25_3.border = Border(bottom=hair, left=thin, right=thin)

    h25_3 = ws3['H25']
    h25_3.value = offer_4_appraisal_contingency_days
    # h25_3.font = Font(bold=True)
    h25_3.alignment = Alignment(horizontal='center')
    h25_3.border = Border(bottom=hair, left=thin, right=thin)

    ws3.merge_cells('C26:C27')
    top_left_home_sale_cont_3 = ws3['C26']
    top_left_home_sale_cont_3.value = 'Home Sale Contingency'
    # top_left_home_sale_cont_3.font = Font(bold=True)
    top_left_home_sale_cont_3.alignment = Alignment(horizontal='left', vertical='center')
    top_left_home_sale_cont_3.border = Border(top=hair, bottom=hair, left=thin)
    c26_3 = ws3['C26']
    c26_3.border = Border(top=hair, left=thin)
    d26_3 = ws3['D26']
    d26_3.border = Border(top=hair)
    c27_3 = ws3['C27']
    c27_3.border = Border(bottom=hair, left=thin)
    d27_3 = ws3['D27']
    d27_3.border = Border(bottom=hair)

    e26_3 = ws3['E26']
    e26_3.value = offer_3_home_sale_contingency_check
    # e26_3.font = Font(bold=True)
    e26_3.alignment = Alignment(horizontal='center')
    e26_3.border = Border(top=hair, left=thin, right=thin)

    f26_3 = ws3['F26']
    f26_3.value = offer_3_home_sale_contingency_check
    # f26_3.font = Font(bold=True)
    f26_3.alignment = Alignment(horizontal='center')
    f26_3.border = Border(top=hair, left=thin, right=thin)

    g26_3 = ws3['G26']
    g26_3.value = offer_3_home_sale_contingency_check
    # g26_3.font = Font(bold=True)
    g26_3.alignment = Alignment(horizontal='center')
    g26_3.border = Border(top=hair, left=thin, right=thin)

    h26_3 = ws3['H26']
    h26_3.value = offer_4_home_sale_contingency_check
    # h26_3.font = Font(bold=True)
    h26_3.alignment = Alignment(horizontal='center')
    h26_3.border = Border(top=hair, left=thin, right=thin)

    e27_3 = ws3['E27']
    e27_3.value = offer_3_home_sale_contingency_days
    # e27_3.font = Font(bold=True)
    e27_3.alignment = Alignment(horizontal='center')
    e27_3.border = Border(bottom=hair, left=thin, right=thin)

    f27_3 = ws3['F27']
    f27_3.value = offer_3_home_sale_contingency_days
    # f27_3.font = Font(bold=True)
    f27_3.alignment = Alignment(horizontal='center')
    f27_3.border = Border(bottom=hair, left=thin, right=thin)

    g27_3 = ws3['G27']
    g27_3.value = offer_3_home_sale_contingency_days
    # g27_3.font = Font(bold=True)
    g27_3.alignment = Alignment(horizontal='center')
    g27_3.border = Border(bottom=hair, left=thin, right=thin)

    h27_3 = ws3['H27']
    h27_3.value = offer_4_home_sale_contingency_days
    # h27_3.font = Font(bold=True)
    h27_3.alignment = Alignment(horizontal='center')
    h27_3.border = Border(bottom=hair, left=thin, right=thin)

    c28_3 = ws3['C28']
    c28_3.value = 'Pre Occupancy Start Date'
    # c28_3.font = Font(bold=True)
    c28_3.alignment = Alignment(horizontal='left', vertical='center')
    c28_3.border = Border(top=hair, bottom=hair, left=thin)

    d28_3 = ws3['D28']
    d28_3.border = Border(top=hair, bottom=hair)

    e28_3 = ws3['E28']
    e28_3.value = offer_3_pre_occupancy_date
    # e28_3.font = Font(bold=True)
    e28_3.alignment = Alignment(horizontal='right')
    e28_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f28_3 = ws3['F28']
    f28_3.value = offer_3_pre_occupancy_date
    # f28_3.font = Font(bold=True)
    f28_3.alignment = Alignment(horizontal='right')
    f28_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g28_3 = ws3['G28']
    g28_3.value = offer_3_pre_occupancy_date
    # g28_3.font = Font(bold=True)
    g28_3.alignment = Alignment(horizontal='right')
    g28_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h28_3 = ws3['H28']
    h28_3.value = offer_4_pre_occupancy_date
    # h28_3.font = Font(bold=True)
    h28_3.alignment = Alignment(horizontal='right')
    h28_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c29_3 = ws3['C29']
    c29_3.value = 'Post Occupancy Thru Date'
    # c29_3.font = Font(bold=True)
    c29_3.alignment = Alignment(horizontal='left', vertical='center')
    c29_3.border = Border(top=hair, bottom=thin, left=thin)

    d29_3 = ws3['D29']
    d29_3.border = Border(top=hair, bottom=thin)

    e29_3 = ws3['E29']
    e29_3.value = offer_3_post_occupancy_date
    # e29_3.font = Font(bold=True)
    e29_3.alignment = Alignment(horizontal='right')
    e29_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f29_3 = ws3['F29']
    f29_3.value = offer_3_post_occupancy_date
    # f29_3.font = Font(bold=True)
    f29_3.alignment = Alignment(horizontal='right')
    f29_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g29_3 = ws3['G29']
    g29_3.value = offer_3_post_occupancy_date
    # g29_3.font = Font(bold=True)
    g29_3.alignment = Alignment(horizontal='right')
    g29_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h29_3 = ws3['H29']
    h29_3.value = offer_4_post_occupancy_date
    # h29_3.font = Font(bold=True)
    h29_3.alignment = Alignment(horizontal='right')
    h29_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    # Housing Related Cost Table
    c31_3 = ws3['C31']
    c31_3.value = 'HOUSING-RELATED COSTS'
    c31_3.font = Font(bold=True)
    c31_3.border = Border(bottom=thin)

    d31_3 = ws3['D31']
    d31_3.value = 'Calculation Description'
    d31_3.font = Font(bold=True)
    d31_3.border = Border(bottom=thin)

    c32_3 = ws3['C32']
    c32_3.value = 'Estimated Payoff - 1st Trust'
    c32_3.border = Border(top=thin, bottom=hair, left=thin)

    d32_3 = ws3['D32']
    d32_3.value = 'Principal Balance of Loan'
    d32_3.border = Border(top=thin, bottom=hair)

    e32_3 = ws3['E32']
    e32_3.value = first_trust
    e32_3.number_format = acct_fmt
    e32_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f32_3 = ws3['F32']
    f32_3.value = first_trust
    f32_3.number_format = acct_fmt
    f32_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g32_3 = ws3['G32']
    g32_3.value = first_trust
    g32_3.number_format = acct_fmt
    g32_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h32_3 = ws3['H32']
    h32_3.value = first_trust
    h32_3.number_format = acct_fmt
    h32_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c33_3 = ws3['C33']
    c33_3.value = 'Estimated Payoff - 2nd Trust'
    c33_3.border = Border(top=hair, bottom=hair, left=thin)

    d33_3 = ws3['D33']
    d33_3.value = 'Principal Balance of Loan'
    d33_3.border = Border(top=hair, bottom=hair)

    e33_3 = ws3['E33']
    e33_3.value = second_trust
    e33_3.number_format = acct_fmt
    e33_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f33_3 = ws3['F33']
    f33_3.value = second_trust
    f33_3.number_format = acct_fmt
    f33_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g33_3 = ws3['G33']
    g33_3.value = second_trust
    g33_3.number_format = acct_fmt
    g33_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h33_3 = ws3['H33']
    h33_3.value = second_trust
    h33_3.number_format = acct_fmt
    h33_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c34_3 = ws3['C34']
    c34_3.value = 'Purchaser Closing Cost / Contract'
    c34_3.border = Border(top=hair, bottom=hair, left=thin)

    d34_3 = ws3['D34']
    d34_3.value = 'Negotiated Into Contract'
    d34_3.border = Border(top=hair, bottom=hair)

    e34_3 = ws3['E34']
    e34_3.value = offer_3_closing_cost_subsidy_amt
    e34_3.number_format = acct_fmt
    e34_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f34_3 = ws3['F34']
    f34_3.value = offer_3_closing_cost_subsidy_amt
    f34_3.number_format = acct_fmt
    f34_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g34_3 = ws3['G34']
    g34_3.value = offer_3_closing_cost_subsidy_amt
    g34_3.number_format = acct_fmt
    g34_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h34_3 = ws3['H34']
    h34_3.value = offer_4_closing_cost_subsidy_amt
    h34_3.number_format = acct_fmt
    h34_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c35_3 = ws3['C35']
    c35_3.value = 'Prorated Taxes / Assessments'
    c35_3.border = Border(top=hair, bottom=hair, left=thin)

    d35_3 = ws3['D35']
    d35_3.value = '1 Year of Taxes Divided by 12 Multiplied by 3'
    d35_3.border = Border(top=hair, bottom=hair)

    e35_3 = ws3['E35']
    e35_3.value = prorated_taxes
    e35_3.number_format = acct_fmt
    e35_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f35_3 = ws3['F35']
    f35_3.value = prorated_taxes
    f35_3.number_format = acct_fmt
    f35_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g35_3 = ws3['G35']
    g35_3.value = prorated_taxes
    g35_3.number_format = acct_fmt
    g35_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h35_3 = ws3['H35']
    h35_3.value = prorated_taxes
    h35_3.number_format = acct_fmt
    h35_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c36_3 = ws3['C36']
    c36_3.value = 'Prorated HOA / Condo Dues'
    c36_3.border = Border(top=hair, bottom=thin, left=thin)

    d36_3 = ws3['D36']
    d36_3.value = '1 Year of Dues Divided by 12 Multiplied by 3'
    d36_3.border = Border(top=hair, bottom=thin)

    e36_3 = ws3['E36']
    e36_3.value = prorated_hoa_condo_fees
    e36_3.number_format = acct_fmt
    e36_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f36_3 = ws3['F36']
    f36_3.value = prorated_hoa_condo_fees
    f36_3.number_format = acct_fmt
    f36_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g36_3 = ws3['G36']
    g36_3.value = prorated_hoa_condo_fees
    g36_3.number_format = acct_fmt
    g36_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h36_3 = ws3['H36']
    h36_3.value = prorated_hoa_condo_fees
    h36_3.number_format = acct_fmt
    h36_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e37_3 = ws3['E37']
    e37_3.value = '=SUM(E32:E36)'
    e37_3.font = Font(bold=True)
    e37_3.number_format = acct_fmt
    e37_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f37_3 = ws3['F37']
    f37_3.value = '=SUM(F32:F36)'
    f37_3.font = Font(bold=True)
    f37_3.number_format = acct_fmt
    f37_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g37_3 = ws3['G37']
    g37_3.value = '=SUM(G32:G36)'
    g37_3.font = Font(bold=True)
    g37_3.number_format = acct_fmt
    g37_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h37_3 = ws3['H37']
    h37_3.value = '=SUM(H32:H36)'
    h37_3.font = Font(bold=True)
    h37_3.number_format = acct_fmt
    h37_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Broker and Financing Costs Table
    c38_3 = ws3['C38']
    c38_3.value = 'BROKERAGE & FINANCING COSTS'
    c38_3.font = Font(bold=True)
    c38_3.border = Border(bottom=thin)

    c39_3 = ws3['C39']
    c39_3.value = 'Listing Company Compensation'
    c39_3.border = Border(top=thin, bottom=hair, left=thin)

    d39_3 = ws3['D39']
    d39_3.value = '% from Listing Agreement * Offer Amount ($)'
    d39_3.border = Border(top=thin, bottom=hair)

    e39_3 = ws3['E39']
    e39_3.value = listing_company_pct * offer_3_amt
    e39_3.number_format = acct_fmt
    e39_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f39_3 = ws3['F39']
    f39_3.value = listing_company_pct * offer_3_amt
    f39_3.number_format = acct_fmt
    f39_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g39_3 = ws3['G39']
    g39_3.value = listing_company_pct * offer_3_amt
    g39_3.number_format = acct_fmt
    g39_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h39_3 = ws3['H39']
    h39_3.value = listing_company_pct * offer_4_amt
    h39_3.number_format = acct_fmt
    h39_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c40_3 = ws3['C40']
    c40_3.value = 'Selling Company Compensation'
    c40_3.border = Border(top=hair, bottom=hair, left=thin)

    d40_3 = ws3['D40']
    d40_3.value = '% from Listing Agreement * Offer Amount ($)'
    d40_3.border = Border(top=hair, bottom=hair)

    e40_3 = ws3['E40']
    e40_3.value = selling_company_pct * offer_3_amt
    e40_3.number_format = acct_fmt
    e40_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f40_3 = ws3['F40']
    f40_3.value = selling_company_pct * offer_3_amt
    f40_3.number_format = acct_fmt
    f40_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g40_3 = ws3['G40']
    g40_3.value = selling_company_pct * offer_3_amt
    g40_3.number_format = acct_fmt
    g40_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h40_3 = ws3['H40']
    h40_3.value = selling_company_pct * offer_4_amt
    h40_3.number_format = acct_fmt
    h40_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c41_3 = ws3['C41']
    c41_3.value = 'Processing Fee'
    c41_3.border = Border(top=hair, bottom=thin, left=thin)

    d41_3 = ws3['D41']
    d41_3.border = Border(top=hair, bottom=thin)

    e41_3 = ws3['E41']
    e41_3.value = processing_fee
    e41_3.number_format = acct_fmt
    e41_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f41_3 = ws3['F41']
    f41_3.value = processing_fee
    f41_3.number_format = acct_fmt
    f41_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g41_3 = ws3['G41']
    g41_3.value = processing_fee
    g41_3.number_format = acct_fmt
    g41_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h41_3 = ws3['H41']
    h41_3.value = processing_fee
    h41_3.number_format = acct_fmt
    h41_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e42_3 = ws3['E42']
    e42_3.value = '=SUM(E39:E41)'
    e42_3.font = Font(bold=True)
    e42_3.number_format = acct_fmt
    e42_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f42_3 = ws3['F42']
    f42_3.value = '=SUM(F39:F41)'
    f42_3.font = Font(bold=True)
    f42_3.number_format = acct_fmt
    f42_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g42_3 = ws3['G42']
    g42_3.value = '=SUM(G39:G41)'
    g42_3.font = Font(bold=True)
    g42_3.number_format = acct_fmt
    g42_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h42_3 = ws3['H42']
    h42_3.value = '=SUM(H39:H41)'
    h42_3.font = Font(bold=True)
    h42_3.number_format = acct_fmt
    h42_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Estimated Closing Costs Table
    c43_3 = ws3['C43']
    c43_3.value = 'ESTIMATED CLOSING COSTS'
    c43_3.font = Font(bold=True)
    c43_3.border = Border(bottom=thin)

    c44_3 = ws3['C44']
    c44_3.value = 'Settlement Fee'
    c44_3.border = Border(top=thin, bottom=hair, left=thin)

    d44_3 = ws3['D44']
    d44_3.value = 'Commonly Used Fee'
    d44_3.border = Border(top=thin, bottom=hair)

    e44_3 = ws3['E44']
    e44_3.value = settlement_fee
    e44_3.number_format = acct_fmt
    e44_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f44_3 = ws3['F44']
    f44_3.value = settlement_fee
    f44_3.number_format = acct_fmt
    f44_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g44_3 = ws3['G44']
    g44_3.value = settlement_fee
    g44_3.number_format = acct_fmt
    g44_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h44_3 = ws3['H44']
    h44_3.value = settlement_fee
    h44_3.number_format = acct_fmt
    h44_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c45_3 = ws3['C45']
    c45_3.value = 'Deed Preparation'
    c45_3.border = Border(top=hair, bottom=hair, left=thin)

    d45_3 = ws3['D45']
    d45_3.value = 'Commonly Used Fee'
    d45_3.border = Border(top=hair, bottom=hair)

    e45_3 = ws3['E45']
    e45_3.value = deed_preparation_fee
    e45_3.number_format = acct_fmt
    e45_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f45_3 = ws3['F45']
    f45_3.value = deed_preparation_fee
    f45_3.number_format = acct_fmt
    f45_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g45_3 = ws3['G45']
    g45_3.value = deed_preparation_fee
    g45_3.number_format = acct_fmt
    g45_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h45_3 = ws3['H45']
    h45_3.value = deed_preparation_fee
    h45_3.number_format = acct_fmt
    h45_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c46_3 = ws3['C46']
    c46_3.value = 'Release of Liens / Trusts'
    c46_3.border = Border(top=hair, bottom=thin, left=thin)

    d46_3 = ws3['D46']
    d46_3.value = 'Commonly Used Fee * Qty of Trusts or Liens'
    d46_3.border = Border(top=hair, bottom=thin)

    e46_3 = ws3['E46']
    e46_3.value = lien_trust_release_fee * lien_trust_release_qty
    e46_3.number_format = acct_fmt
    e46_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f46_3 = ws3['F46']
    f46_3.value = lien_trust_release_fee * lien_trust_release_qty
    f46_3.number_format = acct_fmt
    f46_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g46_3 = ws3['G46']
    g46_3.value = lien_trust_release_fee * lien_trust_release_qty
    g46_3.number_format = acct_fmt
    g46_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h46_3 = ws3['H46']
    h46_3.value = lien_trust_release_fee * lien_trust_release_qty
    h46_3.number_format = acct_fmt
    h46_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e47_3 = ws3['E47']
    e47_3.value = '=SUM(E44:E46)'
    e47_3.font = Font(bold=True)
    e47_3.number_format = acct_fmt
    e47_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f47_3 = ws3['F47']
    f47_3.value = '=SUM(F44:F46)'
    f47_3.font = Font(bold=True)
    f47_3.number_format = acct_fmt
    f47_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g47_3 = ws3['G47']
    g47_3.value = '=SUM(G44:G46)'
    g47_3.font = Font(bold=True)
    g47_3.number_format = acct_fmt
    g47_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h47_3 = ws3['H47']
    h47_3.value = '=SUM(H44:H46)'
    h47_3.font = Font(bold=True)
    h47_3.number_format = acct_fmt
    h47_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Misc Costs Table
    c48_3 = ws3['C48']
    c48_3.value = 'MISCELLANEOUS COSTS'
    c48_3.font = Font(bold=True)
    c48_3.border = Border(bottom=thin)

    c49_3 = ws3['C49']
    c49_3.value = 'Recording Release(s)'
    c49_3.border = Border(top=thin, bottom=hair, left=thin)

    d49_3 = ws3['D49']
    d49_3.value = 'Commonly Used Fee * Qty of Trusts Recorded'
    d49_3.border = Border(top=thin, bottom=hair)

    e49_3 = ws3['E49']
    e49_3.value = recording_fee * recording_trusts_liens_qty
    e49_3.number_format = acct_fmt
    e49_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    f49_3 = ws3['F49']
    f49_3.value = recording_fee * recording_trusts_liens_qty
    f49_3.number_format = acct_fmt
    f49_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    g49_3 = ws3['G49']
    g49_3.value = recording_fee * recording_trusts_liens_qty
    g49_3.number_format = acct_fmt
    g49_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    h49_3 = ws3['H49']
    h49_3.value = recording_fee * recording_trusts_liens_qty
    h49_3.number_format = acct_fmt
    h49_3.border = Border(top=thin, bottom=hair, left=thin, right=thin)

    c50_3 = ws3['C50']
    c50_3.value = 'Grantor\'s Tax'
    c50_3.border = Border(top=hair, bottom=hair, left=thin)

    d50_3 = ws3['D50']
    d50_3.value = '% of Offer Amount ($)'
    d50_3.border = Border(top=hair, bottom=hair)

    e50_3 = ws3['E50']
    e50_3.value = grantors_tax_pct * offer_3_amt
    e50_3.number_format = acct_fmt
    e50_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f50_3 = ws3['F50']
    f50_3.value = grantors_tax_pct * offer_3_amt
    f50_3.number_format = acct_fmt
    f50_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g50_3 = ws3['G50']
    g50_3.value = grantors_tax_pct * offer_3_amt
    g50_3.number_format = acct_fmt
    g50_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h50_3 = ws3['H50']
    h50_3.value = grantors_tax_pct * offer_4_amt
    h50_3.number_format = acct_fmt
    h50_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c51_3 = ws3['C51']
    c51_3.value = 'Congestion Relief Tax'
    c51_3.border = Border(top=hair, bottom=hair, left=thin)

    d51_3 = ws3['D51']
    d51_3.value = '% of Offer Amount ($)'
    d51_3.border = Border(top=hair, bottom=hair)

    e51_3 = ws3['E51']
    e51_3.value = congestion_tax_pct * offer_3_amt
    e51_3.number_format = acct_fmt
    e51_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f51_3 = ws3['F51']
    f51_3.value = congestion_tax_pct * offer_3_amt
    f51_3.number_format = acct_fmt
    f51_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g51_3 = ws3['G51']
    g51_3.value = congestion_tax_pct * offer_3_amt
    g51_3.number_format = acct_fmt
    g51_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h51_3 = ws3['H51']
    h51_3.value = congestion_tax_pct * offer_4_amt
    h51_3.number_format = acct_fmt
    h51_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c52_3 = ws3['C52']
    c52_3.value = 'Pest Inspection'
    c52_3.border = Border(top=hair, bottom=hair, left=thin)

    d52_3 = ws3['D52']
    d52_3.value = 'Commonly Used Fee'
    d52_3.border = Border(top=hair, bottom=hair)

    e52_3 = ws3['E52']
    e52_3.value = pest_inspection_fee
    e52_3.number_format = acct_fmt
    e52_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f52_3 = ws3['F52']
    f52_3.value = pest_inspection_fee
    f52_3.number_format = acct_fmt
    f52_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g52_3 = ws3['G52']
    g52_3.value = pest_inspection_fee
    g52_3.number_format = acct_fmt
    g52_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h52_3 = ws3['H52']
    h52_3.value = pest_inspection_fee
    h52_3.number_format = acct_fmt
    h52_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c53_3 = ws3['C53']
    c53_3.value = 'POA / Condo Disclosures'
    c53_3.border = Border(top=hair, bottom=hair, left=thin)

    d53_3 = ws3['D53']
    d53_3.value = 'Commonly Used Fee'
    d53_3.border = Border(top=hair, bottom=hair)

    e53_3 = ws3['E53']
    e53_3.value = poa_condo_disclosure_fee
    e53_3.number_format = acct_fmt
    e53_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f53_3 = ws3['F53']
    f53_3.value = poa_condo_disclosure_fee
    f53_3.number_format = acct_fmt
    f53_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g53_3 = ws3['G53']
    g53_3.value = poa_condo_disclosure_fee
    g53_3.number_format = acct_fmt
    g53_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h53_3 = ws3['H53']
    h53_3.value = poa_condo_disclosure_fee
    h53_3.number_format = acct_fmt
    h53_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c54_3 = ws3['C54']
    c54_3.value = 'Pre Occupancy Credit to Seller'
    c54_3.border = Border(top=hair, bottom=hair, left=thin)

    d54_3 = ws3['D54']
    d54_3.value = 'Negotiated Into Contract'
    d54_3.border = Border(top=hair, bottom=hair)

    e54_3 = ws3['E54']
    e54_3.value = offer_3_pre_occupancy_credit_amt
    e54_3.number_format = acct_fmt
    e54_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    f54_3 = ws3['F54']
    f54_3.value = offer_3_pre_occupancy_credit_amt
    f54_3.number_format = acct_fmt
    f54_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    g54_3 = ws3['G54']
    g54_3.value = offer_3_pre_occupancy_credit_amt
    g54_3.number_format = acct_fmt
    g54_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    h54_3 = ws3['H54']
    h54_3.value = offer_4_pre_occupancy_credit_amt
    h54_3.number_format = acct_fmt
    h54_3.border = Border(top=hair, bottom=hair, left=thin, right=thin)

    c55_3 = ws3['C55']
    c55_3.value = 'Post Occupancy Cost to Seller'
    c55_3.border = Border(top=hair, bottom=thin, left=thin)

    d55_3 = ws3['D55']
    d55_3.value = 'Negotiated Into Contract'
    d55_3.border = Border(top=hair, bottom=thin)

    e55_3 = ws3['E55']
    e55_3.value = offer_3_post_occupancy_cost_amt
    e55_3.number_format = acct_fmt
    e55_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    f55_3 = ws3['F55']
    f55_3.value = offer_3_post_occupancy_cost_amt
    f55_3.number_format = acct_fmt
    f55_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    g55_3 = ws3['G55']
    g55_3.value = offer_3_post_occupancy_cost_amt
    g55_3.number_format = acct_fmt
    g55_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    h55_3 = ws3['H55']
    h55_3.value = offer_4_post_occupancy_cost_amt
    h55_3.number_format = acct_fmt
    h55_3.border = Border(top=hair, bottom=thin, left=thin, right=thin)

    e56_3 = ws3['E56']
    e56_3.value = '=SUM(E49:E53,E55)-E54'
    e56_3.font = Font(bold=True)
    e56_3.number_format = acct_fmt
    e56_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f56_3 = ws3['F56']
    f56_3.value = '=SUM(F49:F53,F55)-F54'
    f56_3.font = Font(bold=True)
    f56_3.number_format = acct_fmt
    f56_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g56_3 = ws3['G56']
    g56_3.value = '=SUM(G49:G53,G55)-G54'
    g56_3.font = Font(bold=True)
    g56_3.number_format = acct_fmt
    g56_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h56_3 = ws3['H56']
    h56_3.value = '=SUM(H49:H53,H55)-H54'
    h56_3.font = Font(bold=True)
    h56_3.number_format = acct_fmt
    h56_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Subtotals section
    ws3.merge_cells('C58:D58')
    top_left_cell_four_3 = ws3['C58']
    top_left_cell_four_3.value = 'TOTAL ESTIMATED COST OF SETTLEMENT'
    top_left_cell_four_3.font = Font(bold=True)
    top_left_cell_four_3.alignment = Alignment(horizontal='right')

    e58_3 = ws3['E58']
    e58_3.value = '=SUM(E37,E42,E47,E56)'
    e58_3.font = Font(bold=True)
    e58_3.number_format = acct_fmt
    e58_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f58_3 = ws3['F58']
    f58_3.value = '=SUM(F37,F42,F47,F56)'
    f58_3.font = Font(bold=True)
    f58_3.number_format = acct_fmt
    f58_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g58_3 = ws3['G58']
    g58_3.value = '=SUM(G37,G42,G47,G56)'
    g58_3.font = Font(bold=True)
    g58_3.number_format = acct_fmt
    g58_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h58_3 = ws3['H58']
    h58_3.value = '=SUM(H37,H42,H47,H56)'
    h58_3.font = Font(bold=True)
    h58_3.number_format = acct_fmt
    h58_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws3.merge_cells('C59:D59')
    top_left_cell_five_3 = ws3['C59']
    top_left_cell_five_3.value = 'Offer Amount ($)'
    top_left_cell_five_3.font = Font(bold=True)
    top_left_cell_five_3.alignment = Alignment(horizontal='right')

    e59_3 = ws3['E59']
    e59_3.value = offer_3_amt
    e59_3.font = Font(bold=True)
    e59_3.number_format = acct_fmt
    e59_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f59_3 = ws3['F59']
    f59_3.value = offer_3_amt
    f59_3.font = Font(bold=True)
    f59_3.number_format = acct_fmt
    f59_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g59_3 = ws3['G59']
    g59_3.value = offer_3_amt
    g59_3.font = Font(bold=True)
    g59_3.number_format = acct_fmt
    g59_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h59_3 = ws3['H59']
    h59_3.value = offer_4_amt
    h59_3.font = Font(bold=True)
    h59_3.number_format = acct_fmt
    h59_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws3.merge_cells('C60:D60')
    top_left_cell_six_3 = ws3['C60']
    top_left_cell_six_3.value = 'LESS: Total Estimated Cost of Settlement'
    top_left_cell_six_3.font = Font(bold=True)
    top_left_cell_six_3.alignment = Alignment(horizontal='right')

    e60_3 = ws3['E60']
    e60_3.value = '=-SUM(E37,E42,E47,E56)'
    e60_3.font = Font(bold=True)
    e60_3.number_format = acct_fmt
    e60_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f60_3 = ws3['F60']
    f60_3.value = '=-SUM(F37,F42,F47,F56)'
    f60_3.font = Font(bold=True)
    f60_3.number_format = acct_fmt
    f60_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g60_3 = ws3['G60']
    g60_3.value = '=-SUM(G37,G42,G47,G56)'
    g60_3.font = Font(bold=True)
    g60_3.number_format = acct_fmt
    g60_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h60_3 = ws3['H60']
    h60_3.value = '=-SUM(H37,H42,H47,H56)'
    h60_3.font = Font(bold=True)
    h60_3.number_format = acct_fmt
    h60_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws3.merge_cells('C61:D61')
    top_left_cell_seven_3 = ws3['C61']
    top_left_cell_seven_3.value = 'ESTIMATED TOTAL NET PROCEEDS'
    top_left_cell_seven_3.font = Font(bold=True)
    top_left_cell_seven_3.alignment = Alignment(horizontal='right')

    e61_3 = ws3['E61']
    e61_3.value = '=SUM(E59:E60)'
    e61_3.font = Font(bold=True)
    e61_3.number_format = acct_fmt
    e61_3.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    e61_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    f61_3 = ws3['F61']
    f61_3.value = '=SUM(F59:F60)'
    f61_3.font = Font(bold=True)
    f61_3.number_format = acct_fmt
    f61_3.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    f61_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    g61_3 = ws3['G61']
    g61_3.value = '=SUM(G59:G60)'
    g61_3.font = Font(bold=True)
    g61_3.number_format = acct_fmt
    g61_3.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    g61_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    h61_3 = ws3['H61']
    h61_3.value = '=SUM(H59:H60)'
    h61_3.font = Font(bold=True)
    h61_3.number_format = acct_fmt
    h61_3.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    h61_3.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Signature Block
    c63_3 = ws3['C63']
    c63_3.value = 'PREPARED BY:'

    c64_3 = ws3['C64']
    c64_3.value = agent

    e63_3 = ws3['E63']
    e63_3.value = 'SELLER:'

    e64_3 = ws3['E64']
    e64_3.value = seller_name

    # Freedom Logo
    # c53 = ws3['C53']
    freedom_logo = drawing.image.Image('freedom_logo.png')
    freedom_logo.height = 80
    freedom_logo.width = 115
    freedom_logo.anchor = 'C66'
    ws3.add_image(freedom_logo)

    # Disclosure Statement
    ws3.merge_cells('C70:H74')
    top_left_cell_eight_3 = ws3['C70']
    top_left_cell_eight_3.value = '''These estimates are not guaranteed and may not include escrows. Escrow balances are reimbursed by the existing lender. Taxes, rents & association dues are pro-rated at settlement. Under Virginia Law, the seller's proceeds may not be available for up to 2 business days following recording of the deed. Seller acknowledges receipt of this statement.'''
    top_left_cell_eight_3.font = Font(italic=True)
    top_left_cell_eight_3.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

    # hide calculation description column
    ws1.column_dimensions['D'].hidden = True
    ws2.column_dimensions['D'].hidden = True
    ws3.column_dimensions['D'].hidden = True


    # hide unused columns based on number of offers
    if offer_qty == 1:
        ws1.column_dimensions['F'].hidden = True
        ws1.column_dimensions['G'].hidden = True
        ws1.column_dimensions['H'].hidden = True
    elif offer_qty == 2:
        ws1.column_dimensions['G'].hidden = True
        ws1.column_dimensions['H'].hidden = True
    elif offer_qty == 3:
        ws1.column_dimensions['H'].hidden = True
    elif offer_qty == 4:
        ws1.column_dimensions['E'].hidden = False
        ws1.column_dimensions['F'].hidden = False
        ws1.column_dimensions['G'].hidden = False
    elif offer_qty == 5:
        ws2.column_dimensions['F'].hidden = True
        ws2.column_dimensions['G'].hidden = True
        ws2.column_dimensions['H'].hidden = True
    elif offer_qty == 6:
        ws2.column_dimensions['G'].hidden = True
        ws2.column_dimensions['H'].hidden = True
    elif offer_qty == 7:
        ws2.column_dimensions['H'].hidden = True
    elif offer_qty == 8:
        ws2.column_dimensions['E'].hidden = False
        ws2.column_dimensions['F'].hidden = False
        ws2.column_dimensions['G'].hidden = False
    elif offer_qty == 9:
        ws3.column_dimensions['F'].hidden = True
        ws3.column_dimensions['G'].hidden = True
        ws3.column_dimensions['H'].hidden = True
    elif offer_qty == 10:
        ws3.column_dimensions['G'].hidden = True
        ws3.column_dimensions['H'].hidden = True
    elif offer_qty == 11:
        ws3.column_dimensions['H'].hidden = True
    elif offer_qty == 12:
        ws3.column_dimensions['E'].hidden = False
        ws3.column_dimensions['F'].hidden = False
        ws3.column_dimensions['G'].hidden = False

    # wb.save(filename=dest_filename)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        data = BytesIO(tmp.read())

    return data
